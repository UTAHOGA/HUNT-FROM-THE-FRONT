import csv
import json
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[2]
SCRIPT = ROOT / "scripts/build-black-bear-production-documents-2026.py"
SOURCE = ROOT / "pipeline/RAW/hunt_unit_database/2026/csv/2026 Permits/2026 black bear permits reviewed res-nr-total.csv"
XLSX_OUT = ROOT / "processed_data/hard_data_exports/hunt_tables/2026/2026_BLACK_BEAR.xlsx"
PDF_OUT = ROOT / "processed_data/hard_data_exports/hunt_tables/2026/2026_BLACK_BEAR.pdf"
VALIDATION_OUT = ROOT / "processed_data/hard_data_exports/hunt_tables/2026/2026_BLACK_BEAR.validation.json"
BUNDLED_PYTHON = Path(
    "C:/Users/tyler/.cache/codex-runtimes/codex-primary-runtime/dependencies/python/python.exe"
)


EXPECTED_HEADERS = [
    "Hunt Name",
    "Hunt Code",
    "Sex",
    "Species",
    "Weapon",
    "Hunt Type",
    "Season",
    "Res",
    "Non-Res",
    "Total",
]


def run_script() -> None:
    python_exe = BUNDLED_PYTHON if BUNDLED_PYTHON.exists() else Path(sys.executable)
    subprocess.run([str(python_exe), str(SCRIPT)], cwd=ROOT, check=True)


def read_source() -> list[dict[str, str]]:
    with SOURCE.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def test_black_bear_production_documents_are_generated_from_reviewed_export() -> None:
    run_script()

    assert XLSX_OUT.exists()
    assert XLSX_OUT.stat().st_size > 12000
    assert PDF_OUT.exists()
    assert PDF_OUT.stat().st_size > 15000
    assert VALIDATION_OUT.exists()

    validation = json.loads(VALIDATION_OUT.read_text(encoding="utf-8"))
    assert validation["row_count"] == 106
    assert validation["unique_hunt_code_count"] == 106
    assert validation["status_counts"]["FULL_SPLIT"] == 100
    assert validation["status_counts"]["TOTAL_ONLY"] == 1
    assert validation["checked_codes"]["BR7307"]["status"] == "TOTAL_ONLY"
    assert "code reuse" in validation["checked_codes"]["BR7307"]["code_reuse_warning"]


def test_black_bear_xlsx_has_clean_library_headers_and_key_values() -> None:
    run_script()
    wb = load_workbook(XLSX_OUT, data_only=True)
    ws = wb["2026_BLACK_BEAR"]

    headers = [ws.cell(3, col).value for col in range(1, 11)]
    assert headers == EXPECTED_HEADERS
    assert ws.max_row == 109
    assert ws.max_column == 10

    rows = {
        ws.cell(row, 2).value: [ws.cell(row, col).value for col in range(1, 11)]
        for row in range(4, ws.max_row + 1)
    }
    assert len(rows) == 106
    assert rows["BR7004"][7:] == [18, 0, 18]
    assert rows["BR7210"][7:] == [3, 0, 3]
    assert rows["BR7211"][7:] == [26, 3, 29]
    assert rows["BR7317"][7:] == [9, 1, 10]
    assert rows["BR7307"][7:] == [None, None, 4]


def test_black_bear_pdf_contains_expected_codes_and_no_internal_notes_paths() -> None:
    run_script()
    reader = PdfReader(str(PDF_OUT))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)

    assert len(reader.pages) >= 3
    assert "2026 BLACK BEAR" in text
    for code in ["BR7004", "BR7210", "BR7211", "BR7307", "BR7317", "BR7326"]:
        assert code in text
    assert "../notes/DATABASE" not in text
    assert "reviewed 2026 Utah DWR Hunt Planner black bear permits export" in text
    assert "BR7307 is 2026 code reuse" in text


def test_production_xlsx_row_count_matches_reviewed_source() -> None:
    run_script()
    source_rows = read_source()
    wb = load_workbook(XLSX_OUT, data_only=True)
    ws = wb["2026_BLACK_BEAR"]

    source_codes = {row["hunt_code"] for row in source_rows}
    xlsx_codes = {ws.cell(row, 2).value for row in range(4, ws.max_row + 1)}
    assert len(source_rows) == 106
    assert source_codes == xlsx_codes
