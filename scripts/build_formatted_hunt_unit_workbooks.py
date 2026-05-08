from __future__ import annotations

import csv
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REPO = Path(__file__).resolve().parents[1]
CSV_DIR = REPO / "pipeline" / "RAW" / "hunt_unit_database" / "2026" / "csv"
OUT_DIR = REPO / "pipeline" / "RAW" / "hunt_unit_database" / "2026" / "formatted_xlsx"
REPORT_JSON = REPO / "processed_data" / "formatted_hunt_unit_workbooks_20260508.json"
REPORT_MD = REPO / "processed_data" / "formatted_hunt_unit_workbooks_20260508.md"


HEADER_FILL = PatternFill("solid", fgColor="9A4E00")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BODY_FONT = Font(color="2B1C12", size=10)
LINK_FONT = Font(color="0563C1", underline="single", size=10)
THIN = Side(style="thin", color="D8B58E")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def is_source_csv(path: Path) -> bool:
    name = path.name.lower()
    if not name.endswith(".csv"):
        return False
    if ".backup" in name or "backup_before" in name:
        return False
    return True


def safe_sheet_title(name: str) -> str:
    title = re.sub(r"[\[\]\:\*\?\/\\]", " ", name).strip() or "Sheet1"
    return title[:31]


def read_csv(path: Path) -> list[list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    return rows


def target_width(header: str, values: list[str]) -> float:
    lower = header.lower()
    observed = max([len(header), *[len(str(v or "")) for v in values[:250]]], default=len(header))
    if "hunt_code" in lower or "hunt_number" in lower:
        return 13
    if lower in {"sex", "sex_type"}:
        return 14
    if lower == "species":
        return 22
    if "weapon" in lower:
        return 24
    if "hunt_type" in lower:
        return 24
    if "season" in lower:
        return 42
    if "permit" in lower or lower in {"res", "nonres", "total"}:
        return 15
    if "notes" in lower or "other" in lower or "link" in lower or "source" in lower:
        return 42
    if "description" in lower:
        return 58
    if "hunt_name" in lower or lower == "unit":
        return 34
    return max(10, min(observed + 2, 28))


def note_hyperlink(value: str, csv_path: Path) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    if text.startswith("../notes/") or text.startswith("./") or text.endswith(".md"):
        return text.replace("\\", "/")
    return None


def build_workbook(csv_path: Path) -> dict:
    rows = read_csv(csv_path)
    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_title(csv_path.stem)

    for row in rows:
        ws.append(row)

    if not rows:
        out_path = OUT_DIR / f"{csv_path.stem}.xlsx"
        wb.save(out_path)
        return {"source_csv": str(csv_path.relative_to(REPO)).replace("\\", "/"), "xlsx": str(out_path.relative_to(REPO)).replace("\\", "/"), "rows": 0, "columns": 0}

    headers = rows[0]
    max_row = ws.max_row
    max_col = ws.max_column

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.15
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = "1:1"

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 28

    for column_idx, header in enumerate(headers, start=1):
        values = [rows[row_idx][column_idx - 1] if column_idx - 1 < len(rows[row_idx]) else "" for row_idx in range(1, len(rows))]
        ws.column_dimensions[get_column_letter(column_idx)].width = target_width(header, values)

    for row_idx in range(2, max_row + 1):
        max_text = 0
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = BORDER
            link = note_hyperlink(str(cell.value or ""), csv_path)
            if link:
                cell.hyperlink = link
                cell.font = LINK_FONT
                cell.style = "Hyperlink"
            max_text = max(max_text, len(str(cell.value or "")))
        ws.row_dimensions[row_idx].height = 34 if max_text > 80 else 24

    out_path = OUT_DIR / f"{csv_path.stem}.xlsx"
    wb.save(out_path)
    return {
        "source_csv": str(csv_path.relative_to(REPO)).replace("\\", "/"),
        "xlsx": str(out_path.relative_to(REPO)).replace("\\", "/"),
        "rows": max_row - 1,
        "columns": max_col,
        "page_setup": {
            "orientation": "portrait",
            "margins": "narrow",
            "horizontal_centered": True,
            "fit_to_width": 1,
        },
    }


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    REPORT_JSON.parent.mkdir(parents=True, exist_ok=True)

    outputs = []
    for csv_path in sorted(CSV_DIR.iterdir()):
        if is_source_csv(csv_path):
            outputs.append(build_workbook(csv_path))

    REPORT_JSON.write_text(json.dumps({"generated_at": "2026-05-08", "outputs": outputs}, indent=2) + "\n", encoding="utf-8")
    REPORT_MD.write_text(
        "\n".join(
            [
                "# Formatted Hunt Unit Workbooks",
                "",
                "Generated: 2026-05-08",
                "",
                "All workbooks are portrait layout, narrow margins, horizontally centered, fit to one printed page width, frozen header row, filtered headers, wrapped cells, and linked note files where applicable.",
                "",
                "| Source CSV | XLSX | Rows | Columns |",
                "| --- | --- | ---: | ---: |",
                *[f"| {item['source_csv']} | {item['xlsx']} | {item['rows']} | {item['columns']} |" for item in outputs],
                "",
            ]
        ),
        encoding="utf-8",
    )

    print(json.dumps({"ok": True, "workbooks": len(outputs), "output_dir": str(OUT_DIR.relative_to(REPO)).replace("\\", "/"), "report": str(REPORT_MD.relative_to(REPO)).replace("\\", "/")}, indent=2))


if __name__ == "__main__":
    main()
