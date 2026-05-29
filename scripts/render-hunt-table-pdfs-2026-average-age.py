from __future__ import annotations

import csv
import json
import math
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, legal
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas


ROOT = Path(r"C:\Users\tyler\Desktop\GitHub\HUNT-BUILDER")
XLSX_DIR = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "XLXS"
PDF_DIR = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "PDF'S"
AGE_SOURCE = ROOT / "processed_data" / "harvest_age_features_by_hunt_code_latest.csv"
AUDIT_CSV = ROOT / "processed_data" / "audits" / "hunt_tables_2026_pdf_average_age_audit.csv"
AUDIT_JSON = ROOT / "processed_data" / "audits" / "hunt_tables_2026_pdf_average_age_audit.json"

AGE_ALIASES = [
    "Average Age Harvested (previous hunting season)",
    "Average Age Harvested",
    "Average Harvest Age",
    "Avg Age Harvested",
]


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def norm_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())


def norm_code(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", "", clean(value).upper())


def numeric_display(value: Any) -> str:
    try:
        number = float(clean(value))
    except ValueError:
        return ""
    if number <= 0:
        return ""
    return f"{number:.1f}".rstrip("0").rstrip(".")


def load_age_lookup() -> dict[str, dict[str, str]]:
    lookup: dict[str, dict[str, str]] = {}
    with AGE_SOURCE.open(newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            if clean(row.get("review_status")).upper() != "PASS":
                continue
            code = norm_code(row.get("hunt_code"))
            age = numeric_display(row.get("average_harvest_age"))
            if not code or not age:
                continue
            try:
                year = int(float(clean(row.get("reported_hunt_year"))))
            except ValueError:
                year = 0
            prior_year = int(lookup.get(code, {}).get("reported_hunt_year", "0") or 0)
            if year >= prior_year:
                lookup[code] = {
                    "average_harvest_age": age,
                    "reported_hunt_year": str(year) if year else "",
                    "source_file": clean(row.get("source_file")),
                }
    return lookup


def find_header_row(ws) -> int:
    best_row = 1
    best_score = -1
    for row in range(1, min(ws.max_row, 14) + 1):
        values = [clean(cell.value).lower() for cell in ws[row]]
        joined = " | ".join(v for v in values if v)
        score = sum(1 for v in values if v)
        if "hunt_code" in joined or "hunt code" in joined:
            score += 20
        if "average age" in joined:
            score += 10
        if score > best_score:
            best_score = score
            best_row = row
    return best_row


def wrap_cell(text: str, width: float, font_size: float, max_lines: int) -> list[str]:
    text = clean(text)
    if not text:
        return [""]
    max_chars = max(4, int(width / (font_size * 0.5)))
    words = text.split(" ")
    lines: list[str] = []
    line = ""
    for word in words:
        candidate = word if not line else f"{line} {word}"
        if len(candidate) <= max_chars:
            line = candidate
        else:
            if line:
                lines.append(line)
            line = word[:max_chars]
        if len(lines) >= max_lines:
            break
    if line and len(lines) < max_lines:
        lines.append(line)
    if len(lines) == max_lines and len(text) > len(" ".join(lines)):
        lines[-1] = lines[-1][: max(1, max_chars - 3)] + "..."
    return lines or [""]


def column_widths(headers: list[str], usable_width: float) -> list[float]:
    weights = []
    for header in headers:
        key = norm_header(header)
        if key == "huntname":
            weights.append(2.15)
        elif key == "huntcode":
            weights.append(0.9)
        elif key == "season":
            weights.append(2.45)
        elif key == "notes":
            weights.append(1.65)
        elif "percentharvestsuccess" in key or "averageageharvested" in key or "avgdayshunted" in key:
            weights.append(1.18)
        elif "permits2026" in key or key == "harvestprioryear":
            weights.append(0.78)
        else:
            weights.append(1.05)
    total = sum(weights) or 1
    return [usable_width * (weight / total) for weight in weights]


def read_rows(path: Path, age_lookup: dict[str, dict[str, str]]) -> tuple[list[str], list[list[str]], list[dict[str, str]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = find_header_row(ws)
    headers = [clean(ws.cell(header_row, col).value) for col in range(1, ws.max_column + 1)]
    while headers and not headers[-1]:
        headers.pop()

    header_index = {norm_header(header): idx for idx, header in enumerate(headers)}
    code_idx = header_index.get(norm_header("hunt_code")) or header_index.get(norm_header("Hunt Code"))
    age_idx = None
    for alias in AGE_ALIASES:
        if norm_header(alias) in header_index:
            age_idx = header_index[norm_header(alias)]
            break
    prior_year_idx = header_index.get(norm_header("Harvest Prior Year"))

    rows: list[list[str]] = []
    audit_rows: list[dict[str, str]] = []
    for row_number in range(header_row + 1, ws.max_row + 1):
        values = [clean(ws.cell(row_number, col).value) for col in range(1, len(headers) + 1)]
        if not any(values):
            continue
        code = norm_code(values[code_idx]) if code_idx is not None and code_idx < len(values) else ""
        hard = age_lookup.get(code, {})
        existing_age = values[age_idx] if age_idx is not None and age_idx < len(values) else ""
        status = "AGE_BLANK_NO_PASS_NUMERIC_HARD_DATA"
        if hard and age_idx is not None:
            values[age_idx] = hard["average_harvest_age"]
            status = "PDF_AGE_RENDERED_FROM_PASS_HARD_DATA"
            if prior_year_idx is not None and prior_year_idx < len(values) and not values[prior_year_idx]:
                values[prior_year_idx] = hard.get("reported_hunt_year", "")
        elif existing_age:
            status = "PDF_EXISTING_AGE_NO_PASS_HARD_DATA"
        rows.append(values)
        audit_rows.append(
            {
                "workbook": path.name,
                "pdf": f"{path.stem}.pdf",
                "row_number": str(row_number),
                "hunt_code": code,
                "age_value_rendered": values[age_idx] if age_idx is not None and age_idx < len(values) else "",
                "hard_data_age": hard.get("average_harvest_age", ""),
                "hard_data_reported_hunt_year": hard.get("reported_hunt_year", ""),
                "status": status,
            }
        )
    wb.close()
    return headers, rows, audit_rows


def draw_pdf(path: Path, title: str, headers: list[str], rows: list[list[str]]) -> None:
    page_width, page_height = landscape(legal)
    margin_x = 0.18 * inch
    margin_y = 0.22 * inch
    usable_width = page_width - (2 * margin_x)
    y_start = page_height - margin_y
    col_widths = column_widths(headers, usable_width)
    header_height = 0.42 * inch
    row_height = 0.34 * inch
    rows_per_page = max(1, math.floor((page_height - (2 * margin_y) - header_height - 0.22 * inch) / row_height))

    c = canvas.Canvas(str(path), pagesize=landscape(legal))

    def draw_header() -> None:
        c.setFont("Helvetica-Bold", 10)
        c.setFillColor(colors.HexColor("#2F1B0F"))
        c.drawCentredString(page_width / 2, y_start, title)
        y = y_start - 0.2 * inch
        x = margin_x
        c.setFillColor(colors.HexColor("#5B301B"))
        c.rect(margin_x, y - header_height, usable_width, header_height, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 5.9)
        for idx, header in enumerate(headers):
            width = col_widths[idx]
            for line_idx, line in enumerate(wrap_cell(header, width - 4, 5.9, 3)[:3]):
                c.drawString(x + 2, y - 9 - (line_idx * 7), line)
            c.setStrokeColor(colors.HexColor("#C58F61"))
            c.rect(x, y - header_height, width, header_height, stroke=1, fill=0)
            x += width

    row_index = 0
    page = 0
    while row_index < len(rows) or page == 0:
        page += 1
        draw_header()
        y = y_start - 0.2 * inch - header_height
        page_rows = rows[row_index : row_index + rows_per_page]
        c.setFont("Helvetica", 5.8)
        for local_idx, row in enumerate(page_rows):
            y -= row_height
            c.setFillColor(colors.HexColor("#F7F1E8") if local_idx % 2 == 0 else colors.HexColor("#EFE5D7"))
            c.rect(margin_x, y, usable_width, row_height, fill=1, stroke=0)
            x = margin_x
            c.setFillColor(colors.black)
            for idx, value in enumerate(row):
                width = col_widths[idx]
                for line_idx, line in enumerate(wrap_cell(value, width - 4, 5.8, 2)[:2]):
                    c.drawString(x + 2, y + row_height - 9 - (line_idx * 7), line)
                c.setStrokeColor(colors.HexColor("#C58F61"))
                c.rect(x, y, width, row_height, stroke=1, fill=0)
                x += width
        row_index += len(page_rows)
        c.setFont("Helvetica", 7)
        c.setFillColor(colors.HexColor("#5B3A25"))
        c.drawRightString(page_width - margin_x, margin_y * 0.55, f"Page {page}")
        if row_index < len(rows):
            c.showPage()
    c.save()


def main() -> None:
    age_lookup = load_age_lookup()
    PDF_DIR.mkdir(parents=True, exist_ok=True)
    audit_rows: list[dict[str, str]] = []
    status_counts: Counter[str] = Counter()
    file_counts: dict[str, Counter[str]] = defaultdict(Counter)

    files = sorted(path for path in XLSX_DIR.glob("*.xlsx") if not path.name.startswith("~$"))
    for workbook in files:
        headers, rows, workbook_audit = read_rows(workbook, age_lookup)
        pdf = PDF_DIR / f"{workbook.stem}.pdf"
        draw_pdf(pdf, workbook.stem, headers, rows)
        audit_rows.extend(workbook_audit)
        for row in workbook_audit:
            status_counts[row["status"]] += 1
            file_counts[workbook.name][row["status"]] += 1
        print(f"OK {pdf.name} rows={len(rows)}")

    AUDIT_CSV.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "workbook",
        "pdf",
        "row_number",
        "hunt_code",
        "age_value_rendered",
        "hard_data_age",
        "hard_data_reported_hunt_year",
        "status",
    ]
    with AUDIT_CSV.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(audit_rows)

    summary = {
        "created_at": datetime.now(timezone.utc).isoformat(),
        "mode": "PDF_ONLY_NO_XLSX_WRITES",
        "xlsx_files_read": len(files),
        "pdf_files_written": len(files),
        "rows_audited": len(audit_rows),
        "pass_numeric_hard_age_codes": len(age_lookup),
        "status_counts": dict(status_counts),
        "file_counts": {name: dict(counts) for name, counts in sorted(file_counts.items())},
        "outputs": {
            "pdf_dir": str(PDF_DIR.relative_to(ROOT)),
            "audit_csv": str(AUDIT_CSV.relative_to(ROOT)),
            "audit_json": str(AUDIT_JSON.relative_to(ROOT)),
        },
    }
    AUDIT_JSON.write_text(json.dumps(summary, indent=2) + "\n", encoding="utf-8")

    print(f"pdf_files_written={summary['pdf_files_written']}")
    print(f"rows_audited={summary['rows_audited']}")
    print(f"pdf_age_rendered_from_pass_hard_data={status_counts['PDF_AGE_RENDERED_FROM_PASS_HARD_DATA']}")
    print(f"age_blank_no_pass_numeric_hard_data={status_counts['AGE_BLANK_NO_PASS_NUMERIC_HARD_DATA']}")
    print(f"audit_json={AUDIT_JSON}")


if __name__ == "__main__":
    main()
