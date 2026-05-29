from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
import re
from typing import List, Tuple

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

SOURCE_DIR = Path(r"C:\Users\tyler\Desktop\GitHub\HUNT-BUILDER\pipeline\RAW\hunt_unit_database\2026\formatted_xlsx")
OUTPUT_DIR = SOURCE_DIR.parent / "formatted_pdf"

HEADER_KEYWORDS = {"hunt_code", "hunt name", "hunt_name", "species", "weapon", "hunt_type"}


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def pretty_header(value: str) -> str:
    text = value.strip().replace("_", " ")
    text = re.sub(r"\s+", " ", text)
    if text.lower().startswith("permits 2026"):
        return text.upper().replace("PERMITS 2026", "2026 PERMITS")
    return text.title()


def find_header_row(sheet) -> int:
    best_row = 1
    best_score = -1
    max_probe = min(sheet.max_row, 12)

    for r in range(1, max_probe + 1):
        row = [normalize_text(c) for c in next(sheet.iter_rows(min_row=r, max_row=r, values_only=True))]
        non_empty = [c for c in row if c]
        if not non_empty:
            continue
        lowered = " | ".join(c.lower() for c in non_empty)
        keyword_hits = sum(1 for k in HEADER_KEYWORDS if k in lowered)
        score = keyword_hits * 10 + len(non_empty)
        if score > best_score:
            best_score = score
            best_row = r

    return best_row


def extract_table(sheet) -> Tuple[str, List[str], List[List[str]]]:
    header_row = find_header_row(sheet)

    title = normalize_text(sheet.cell(row=1, column=1).value)
    if not title:
        title = sheet.title

    header_values = [normalize_text(c) for c in next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]

    # Determine active columns: header populated or data populated below.
    max_col = sheet.max_column
    active_cols = []
    for c in range(1, max_col + 1):
        h = normalize_text(sheet.cell(row=header_row, column=c).value)
        has_data = False
        scan_end = min(sheet.max_row, header_row + 80)
        for r in range(header_row + 1, scan_end + 1):
            if normalize_text(sheet.cell(row=r, column=c).value):
                has_data = True
                break
        if h or has_data:
            active_cols.append(c)

    if not active_cols:
        active_cols = list(range(1, max_col + 1))

    headers = [pretty_header(normalize_text(sheet.cell(row=header_row, column=c).value) or f"Column {i+1}") for i, c in enumerate(active_cols)]

    rows: List[List[str]] = []
    for r in range(header_row + 1, sheet.max_row + 1):
        values = [normalize_text(sheet.cell(row=r, column=c).value) for c in active_cols]
        if not any(values):
            continue
        rows.append(values)

    return title, headers, rows


def fit_column_widths(headers: List[str], rows: List[List[str]], available_width: float) -> List[float]:
    weights = []
    for i, h in enumerate(headers):
        max_len = len(h)
        sample = rows[:400]
        for row in sample:
            max_len = max(max_len, len(row[i]))
        weight = min(max(max_len, 8), 44)
        # favor key columns
        hl = h.lower()
        if "hunt name" in hl:
            weight = max(weight, 24)
        elif "season" in hl:
            weight = max(weight, 22)
        elif "notes" in hl:
            weight = max(weight, 20)
        elif "code" in hl:
            weight = max(weight, 12)
        weights.append(weight)

    total = sum(weights)
    widths = [(w / total) * available_width for w in weights]

    # Clamp then renormalize to keep balanced.
    min_w = 0.65 * inch
    max_w = 2.4 * inch
    widths = [max(min_w, min(max_w, w)) for w in widths]
    scale = available_width / sum(widths)
    widths = [w * scale for w in widths]
    return widths


def make_paragraph(text: str, style: ParagraphStyle) -> Paragraph:
    text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    return Paragraph(text if text else "&nbsp;", style)


def render_pdf(xlsx_path: Path, out_path: Path) -> Tuple[int, int]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    title, headers, rows = extract_table(ws)
    wb.close()

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitlePolished",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=14,
        textColor=colors.HexColor("#2f1b0f"),
        spaceAfter=4,
    )
    meta_style = ParagraphStyle(
        "MetaPolished",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.5,
        textColor=colors.HexColor("#5b3a25"),
        spaceAfter=6,
    )
    cell_style = ParagraphStyle(
        "CellPolished",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.8,
        leading=9.4,
    )
    head_style = ParagraphStyle(
        "HeadPolished",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8.3,
        textColor=colors.white,
        leading=9.8,
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(out_path),
        pagesize=landscape(letter),
        leftMargin=0.35 * inch,
        rightMargin=0.35 * inch,
        topMargin=0.35 * inch,
        bottomMargin=0.35 * inch,
    )

    available_width = doc.width
    col_widths = fit_column_widths(headers, rows, available_width)

    table_data: List[List[Paragraph]] = []
    table_data.append([make_paragraph(h, head_style) for h in headers])

    for row in rows:
        table_data.append([make_paragraph(v, cell_style) for v in row])

    table = Table(table_data, repeatRows=1, colWidths=col_widths, hAlign="LEFT")

    tbl_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4a2e1f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#c7b59f")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#fbf6ef"), colors.HexColor("#f3e8da")]),
    ])
    table.setStyle(tbl_style)

    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    meta = f"Source workbook: {xlsx_path.name} | Rows: {len(rows)} | Generated: {generated}"

    story = [
        Paragraph(title, title_style),
        Paragraph(meta, meta_style),
        Spacer(1, 0.08 * inch),
        table,
    ]

    doc.build(story)
    return len(headers), len(rows)


def main() -> None:
    if not SOURCE_DIR.exists():
        raise SystemExit(f"Missing source directory: {SOURCE_DIR}")

    files = sorted([p for p in SOURCE_DIR.glob("2026*.xlsx") if p.is_file()])
    if not files:
        raise SystemExit("No 2026*.xlsx files found to convert")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    converted = 0
    for xlsx in files:
        pdf_name = xlsx.with_suffix('.pdf').name
        out_pdf = OUTPUT_DIR / pdf_name
        try:
            cols, rows = render_pdf(xlsx, out_pdf)
            print(f"OK  {xlsx.name} -> {out_pdf.name} (cols={cols}, rows={rows})")
            converted += 1
        except Exception as exc:
            print(f"ERR {xlsx.name}: {exc}")

    print(f"DONE converted={converted} total={len(files)} output={OUTPUT_DIR}")


if __name__ == "__main__":
    main()
