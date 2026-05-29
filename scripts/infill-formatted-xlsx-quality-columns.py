from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook
import csv
import re

ROOT = Path(r"C:\Users\tyler\Desktop\GitHub\HUNT-BUILDER")
XLSX_DIR = ROOT / "pipeline" / "RAW" / "hunt_unit_database" / "2026" / "formatted_xlsx"
MODEL_FILE = ROOT / "data_model" / "harvest_quality" / "harvest_feature_model_by_hunt_code_2026.csv"
FALLBACK_FILE = ROOT / "processed_data" / "harvest_master.csv"
AUDIT_FILE = ROOT / "processed_data" / "audits" / "formatted_xlsx_quality_column_infill_audit.csv"

NEW_COLS = [
    "percent_harvest_success_prior_year",
    "avg_days_hunted_prior_year",
    "avg_harvested_age_prior_year",
    "hunter_satisfaction_prior_year",
]


def norm_code(v: str) -> str:
    s = (v or "").strip().upper()
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[^A-Z0-9]", "", s)
    return s


def fmt_num(v: str, *, pct=False) -> str:
    if v is None:
        return "Not available"
    s = str(v).strip()
    if not s:
        return "Not available"
    try:
        n = float(s)
    except ValueError:
        return s
    if pct:
        if n.is_integer():
            return f"{int(n)}%"
        return f"{n:.1f}%"
    if n.is_integer():
        return str(int(n))
    return f"{n:.1f}"


def load_lookup() -> dict:
    out = {}
    with MODEL_FILE.open("r", encoding="utf-8-sig", newline="") as f:
        r = csv.DictReader(f)
        for row in r:
            code = norm_code(row.get("hunt_code", ""))
            if not code:
                continue
            out[code] = {
                "percent_harvest_success_prior_year": fmt_num(row.get("harvest_success_recent"), pct=True),
                "avg_days_hunted_prior_year": fmt_num(row.get("hunter_effort_days_recent")),
                "avg_harvested_age_prior_year": fmt_num(row.get("average_age_recent")),
                "hunter_satisfaction_prior_year": fmt_num(row.get("hunter_satisfaction_recent")),
                "source": "harvest_feature_model_by_hunt_code_2026",
            }

    # fallback fill only where still missing
    with FALLBACK_FILE.open("r", encoding="utf-8-sig", newline="") as f:
        r = csv.DictReader(f)
        by_code = {}
        for row in r:
            code = norm_code(row.get("hunt_code", ""))
            if not code:
                continue
            y = row.get("year", "")
            try:
                yr = int(float(str(y).strip()))
            except Exception:
                yr = -1
            prev = by_code.get(code)
            if prev is None or yr > prev[0]:
                by_code[code] = (yr, row)

    for code, (_, row) in by_code.items():
        base = out.get(code, {
            "percent_harvest_success_prior_year": "Not available",
            "avg_days_hunted_prior_year": "Not available",
            "avg_harvested_age_prior_year": "Not available",
            "hunter_satisfaction_prior_year": "Not available",
            "source": "harvest_master_fallback",
        })

        if base["percent_harvest_success_prior_year"] == "Not available":
            base["percent_harvest_success_prior_year"] = fmt_num(row.get("percent_success"), pct=True)
        if base["avg_days_hunted_prior_year"] == "Not available":
            base["avg_days_hunted_prior_year"] = fmt_num(row.get("avg_days"))
        if base["hunter_satisfaction_prior_year"] == "Not available":
            base["hunter_satisfaction_prior_year"] = fmt_num(row.get("satisfaction"))

        out[code] = base

    return out


def find_header_row(ws) -> int:
    best_row, best_score = 1, -1
    max_probe = min(12, ws.max_row)
    for r in range(1, max_probe + 1):
        vals = [str(c.value).strip() if c.value is not None else "" for c in ws[r]]
        non = [v for v in vals if v]
        if not non:
            continue
        low = " | ".join(v.lower() for v in non)
        score = (10 if "hunt_code" in low else 0) + (7 if "hunt_name" in low else 0) + len(non)
        if score > best_score:
            best_score = score
            best_row = r
    return best_row


def ensure_columns(ws, header_row: int) -> dict:
    max_col = ws.max_column
    header_map = {}
    for c in range(1, max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        key = str(v).strip().lower()
        header_map[key] = c

    insert_pos = ws.max_column + 1
    # try to place before notes if present
    for k, c in header_map.items():
        if k == "notes" or k == "note":
            insert_pos = c
            break

    col_positions = {}
    for col_name in NEW_COLS:
        k = col_name.lower()
        if k in header_map:
            col_positions[col_name] = header_map[k]
        else:
            ws.insert_cols(insert_pos, 1)
            ws.cell(row=header_row, column=insert_pos).value = col_name
            col_positions[col_name] = insert_pos
            insert_pos += 1
            # shift any existing tracked positions to the right if needed
            for existing in list(col_positions):
                if existing != col_name and col_positions[existing] >= col_positions[col_name]:
                    col_positions[existing] += 1

    return col_positions


def process_workbook(path: Path, lookup: dict) -> dict:
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    header_row = find_header_row(ws)

    # locate hunt_code column
    hunt_code_col = None
    for c in range(1, ws.max_column + 1):
        hv = ws.cell(row=header_row, column=c).value
        if hv is None:
            continue
        if str(hv).strip().lower() == "hunt_code":
            hunt_code_col = c
            break

    if hunt_code_col is None:
        wb.close()
        return {"file": path.name, "updated": False, "reason": "NO_HUNT_CODE_HEADER", "rows": 0, "matched": 0}

    col_pos = ensure_columns(ws, header_row)

    rows = 0
    matched = 0
    for r in range(header_row + 1, ws.max_row + 1):
        raw_code = ws.cell(row=r, column=hunt_code_col).value
        code = norm_code(str(raw_code) if raw_code is not None else "")
        if not code:
            continue
        rows += 1
        m = lookup.get(code)
        if m:
            matched += 1
        for col_name in NEW_COLS:
            val = (m or {}).get(col_name, "Not available")
            ws.cell(row=r, column=col_pos[col_name]).value = val

    wb.save(path)
    wb.close()

    return {"file": path.name, "updated": True, "reason": "OK", "rows": rows, "matched": matched}


def main() -> None:
    lookup = load_lookup()
    files = sorted([p for p in XLSX_DIR.glob("2026*.xlsx") if p.is_file()])
    audit_rows = []

    for p in files:
        result = process_workbook(p, lookup)
        audit_rows.append(result)
        print(f"{result['file']}: {result['reason']} rows={result['rows']} matched={result['matched']}")

    AUDIT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with AUDIT_FILE.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["file", "updated", "reason", "rows", "matched"])
        w.writeheader()
        w.writerows(audit_rows)

    total_rows = sum(r["rows"] for r in audit_rows)
    total_matched = sum(r["matched"] for r in audit_rows)
    print(f"DONE files={len(audit_rows)} rows={total_rows} matched={total_matched} audit={AUDIT_FILE}")


if __name__ == "__main__":
    main()
