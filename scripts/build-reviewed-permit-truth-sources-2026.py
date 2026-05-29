from __future__ import annotations

import csv
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATABASE_CSV = ROOT / "pipeline/RAW/hunt_unit_database/2026/csv/DATABASE.csv"
PERMIT_DIR = ROOT / "pipeline/RAW/hunt_unit_database/2026/csv/2026 Permits"
PUBLIC_XLSX_DIR = ROOT / "processed_data/hard_data_exports/hunt_tables/2026/XLXS"
AUDIT_DIR = ROOT / "processed_data/audits"
MASTER_REVIEWED_PERMIT_CSV = PERMIT_DIR / "2026 reviewed permit truth master.csv"
SUPERSEDED_FRAGMENT_MANIFEST = PERMIT_DIR / "2026 superseded permit fragments manifest.csv"


CANONICAL_COLUMNS = [
    "hunt_name",
    "hunt_code",
    "boundary_id",
    "hunt_code_mapping_status",
    "boundary_id_mapping_status",
    "candidate_hunt_code",
    "candidate_boundary_id",
    "sex_type",
    "species",
    "weapon",
    "hunt_type",
    "season",
    "permits_2026_res",
    "permits_2026_nr",
    "permits_2026_total",
    "permit_count_status",
    "source_file",
    "source_row_number",
]

MASTER_COLUMNS = ["reviewed_family"] + CANONICAL_COLUMNS


SUPERSEDED_FRAGMENT_FILES = [
    "2026 black bear permits.csv",
    "black bear.csv",
    "2026 l.e. elk.csv",
    "2026 rocky mountain bighorn user verified.csv",
    "elk antlerless private lands only EA.csv",
    "elk antlerless private lands.csv",
    "2026 buck deer limited entry reviewed res-nr-total.csv",
    "2026 buck deer private land reviewed res-nr-total.csv",
]


REVIEWED_DATABASE_EXCLUSIONS = [
    {
        "hunt_code": "DA1044",
        "hunt_name": "Myton",
        "species": "Deer",
        "sex_type": "Antlerless",
        "review_status": "HOLD_HISTORICAL_NO_CURRENT_DWR_ROW",
        "reason": (
            "User reviewed Utah DWR Hunt Planner and found no current DA1044 result and no same/similar "
            "current hunt name to crosswalk; do not promote without a current pasted DWR source row."
        ),
    },
]


FAMILY_CONFIGS = [
    {
        "family": "BLACK_BEAR",
        "output": "2026 black bear permits reviewed res-nr-total.csv",
        "sources": ["2026 BLACK BEAR DRAW.xlsx"],
        "code_prefixes": ("BR",),
        "species": ("Black Bear",),
    },
    {
        "family": "BISON_HUNTERS_CHOICE",
        "output": "2026 bison reviewed res-nr-total.csv",
        "source_label": "USER_PASTED_DWR_HUNT_PLANNER_BISON_BI_2026",
        "inline_tsv": """
Bison - Statewide Permit	BI1000	Hunters Choice	Bison	Any Legal Weapon	Statewide	Aug 1, 2026 - Jan 31, 2027
Antelope Island	BI6500	Hunters Choice	Bison	Any Legal Weapon	Once-in-a-lifetime	Dec 07 2026 - Dec 09 2026	Res: 2
NonRes: 0
Henry Mtns	BI6503	Hunters Choice	Bison	Any Legal Weapon	Once-in-a-lifetime	Oct 28 2026 - Nov 8 2026	Res: 4
NonRes: 2
Henry Mtns	BI6504	Hunters Choice	Bison	Any Legal Weapon	Once-in-a-lifetime	Nov 11 2026 - Nov 22 2026	Res: 5
NonRes: 2
Henry Mtns	BI6509	Hunters Choice	Bison	Archery	Once-in-a-lifetime	Jan 18 - 31, 2027	Res: 5
NonRes: 0
Henry Mtns	BI6516	Hunters Choice	Bison	Any Legal Weapon	Once-in-a-lifetime	Nov 25 2026 - Dec 6 2026	Res: 6
NonRes: 0
Nine Mile	BI6527	Hunters Choice	Bison	Any Legal Weapon	General Season	Aug 1, 2025 - Jan 31, 2026
Book Cliffs, Little Creek/South	BI6528	Hunters Choice	Bison	Archery	Once-in-a-lifetime	Aug 15 2026 - Sept 11 2026	Res: 6
NonRes: 0
Book Cliffs, Little Creek/South	BI6531	Hunters Choice	Bison	Any Legal Weapon	Once-in-a-lifetime	Sept 21 2026 - Oct 02 2026	Res: 10
NonRes: 0
Book Cliffs, Bitter Creek	BI6532	Hunters Choice	Bison	Archery	Once-in-a-lifetime	Aug 15 2026 - Sept 11 2026	Res: 6
NonRes: 0
Book Cliffs, Bitter Creek	BI6534	Hunters Choice	Bison	Any Legal Weapon	Once-in-a-lifetime	Sept 12 2026 - Sept 20 2026	Res: 10
NonRes: 0
Book Cliffs, Bitter Creek	BI6535	Hunters Choice	Bison	Any Legal Weapon	Once-in-a-lifetime	Oct 03 2026 - Oct 15 2026	Res: 8
NonRes: 0
Book Cliffs, Little Creek/South	BI6537	Hunters Choice	Bison	Any Legal Weapon	Once-in-a-lifetime	Nov 07 2026 - Jan 31 2027	Res: 8
NonRes: 7
Nine Mile, Private Lands Only	BI6538	Hunters Choice	Bison	Any Legal Weapon	General Season	Aug 1, 2025 - Jan 31, 2026 | Harvest survey due by Feb 15, 2026 Visit: http://wildlife.utah.gov/harvest to submit your survey
        """,
        "code_prefixes": ("BI",),
        "species": ("Bison",),
        "sex_type": ("Hunters Choice",),
    },
    {
        "family": "BISON_COW_ONLY",
        "output": "2026 bison cow only reviewed res-nr-total.csv",
        "source_label": "USER_PASTED_DWR_HUNT_PLANNER_BISON_COW_ONLY_BI_2026",
        "inline_tsv": """
Henry Mtns	BI6505	Cow Only	Bison	Any Legal Weapon	Once-in-a-lifetime	Dec 9 2026 - Dec 20 2026	Res: 8
NonRes: 3
Henry Mtns	BI6506	Cow Only	Bison	Any Legal Weapon	Once-in-a-lifetime	Dec 23, 2026 - Jan 3, 2027	Res: 11
NonRes: 0
Book Cliffs, Little Creek/South	BI6529	Cow Only	Bison	Any Legal Weapon	Once-in-a-lifetime	Oct 17 2026 - Oct 27 2026	Res: 8
NonRes: 0
Book Cliffs, Bitter Creek	BI6536	Cow Only	Bison	Any Legal Weapon	Once-in-a-lifetime	Nov 14 2026 - Nov 29 2026	Res: 15
NonRes: 0
Henry Mtns	BI6539	Cow Only	Bison	Any Legal Weapon	Once-in-a-lifetime	Jan 6 - 17, 2027	Res: 11
NonRes: 0
        """,
        "code_prefixes": ("BI",),
        "species": ("Bison",),
        "sex_type": ("Cow Only",),
    },
    {
        "family": "DEER_ANTLERLESS",
        "output": "2026 deer antlerless reviewed res-nr-total.csv",
        "source_label": "USER_PASTED_DWR_HUNT_PLANNER_DEER_ANTLERLESS_DA_2026",
        "inline_tsv": """
Box Elder, West Bear River	DA1001	Antlerless	Deer	Archery, Muzzleloader, Shotgun	General Season	Aug 18 2025 - Aug 30 2025	Total: 30
Box Elder, West Bear River	DA1002	Antlerless	Deer	Archery, Muzzleloader, Shotgun	General Season	Nov 24 2025 - Nov 29 2025	Total: 30
Monroe/Plateau, Sevier Valley	DA1003	Antlerless	Deer	Archery, Muzzleloader, Shotgun	General Season	Sep 01 2025 - Sep 30 2025	Total: 15
Pine Valley, Enterprise	DA1009	Antlerless	Deer	Any Legal Weapon	General Season	Aug 01 2025 - Aug 15 2025	Total: 25
George Creek CWMU	DA1011	Antlerless	Deer	Any Legal Weapon	CWMU	Contact operator for 2025 season dates	Res: 5
NonRes: 0
Junction Valley CWMU	DA1012	Antlerless	Deer	Any Legal Weapon	CWMU	Contact operator for 2025 season dates	Res: 15
NonRes: 0
Mt Carmel CWMU	DA1013	Antlerless	Deer	Any Legal Weapon	CWMU	Contact operator for 2025 season dates	Res: 5
NonRes: 0
Pine Valley, New Harmony	DA1018	Antlerless	Deer	Archery, Muzzleloader, Shotgun	General Season	Aug 01 2025 - Aug 15 2025	Total: 20
East Canyon, Davis-North Salt Lake	DA1027	Antlerless	Deer	Any Legal Weapon	General Season	Aug 01 2025 - Sep 15 2025	Total: 30
Nine Mile, Price River Valley	DA1030	Antlerless	Deer	Archery	General Season	Aug 16 2025 - Sep 12 2025	Total: 25
Oquirrh-Stansbury, Settlement Cyn	DA1031	Antlerless	Deer	Archery	General Season	Aug 16 2025 - Sep 12 2025	Total: 20
San Juan, Monticello	DA1033	Antlerless	Deer	Archery	General Season	Aug 01 2025 - Sep 14 2025	Total: 5
Nine Mile, Green River Valley	DA1041	Antlerless	Deer	Any Legal Weapon	General Season	Oct 04 2025 - Oct 26 2025	Total: 15
San Juan, Monticello	DA1042	Antlerless	Deer	Archery	General Season	Nov 24 2025 - Dec 31 2025	Total: 5
Fillmore, Oak City	DA1045	Antlerless	Deer	Archery	General Season	Aug 01 2025 - Sep 04 2025	Total: 20
Nine Mile, Price River Valley	DA1046	Antlerless	Deer	Archery, Muzzleloader, Shotgun	General Season	Sep 13 2025 - Oct 12 2025	Total: 25
Vernal-Ouray Valley	DA1047	Antlerless	Deer	Any Legal Weapon	General Season	Sep 27 2025 - Dec 12 2025	Total: 25
Fillmore City	DA1048	Antlerless	Deer	Archery	General Season	Aug 01 2025 - Sep 04 2025	Total: 15
Vernal, Ashley Valley	DA1049	Antlerless	Deer	Archery	General Season	Aug 16 2025 - Dec 12 2025	Total: 35
The Rose of Snowville CWMU	DA1050	Antlerless	Deer	Any Legal Weapon	CWMU	Contact operator for 2025 season dates	Res: 8
NonRes: 0
        """,
        "code_prefixes": ("DA",),
        "species": ("Deer",),
        "sex_type": ("Antlerless",),
    },
    {
        "family": "COUGAR_EITHER_SEX",
        "output": "2026 cougar reviewed res-nr-total.csv",
        "source_label": "USER_PASTED_DWR_HUNT_PLANNER_COUGAR_CG_2026",
        "inline_tsv": """
Cougar - Statewide	CG9999	Either Sex	Cougar	Any Legal Weapon	Statewide	unlimited permits.
        """,
        "code_prefixes": ("CG",),
        "species": ("Cougar",),
        "sex_type": ("Either Sex",),
    },
    {
        "family": "ELK_EXTENDED_ARCHERY",
        "output": "2026 elk extended archery reviewed res-nr-total.csv",
        "source_label": "USER_PASTED_DWR_HUNT_PLANNER_ELK_EXTENDED_ARCHERY_EX_2026",
        "inline_tsv": """
Elk Extended Archery	EX1000	Hunters Choice	Elk	Archery	Extended Archery	Aug 16 - Dec 15, 2025
        """,
        "code_prefixes": ("EX",),
        "species": ("Elk",),
        "sex_type": ("Hunters Choice",),
    },
    {
        "family": "BUCK_DEER",
        "output": "2026 buck deer all reviewed res-nr-total.csv",
        "sources": ["2026 DEER BUCK DRAW.xlsx"],
        "code_prefixes": ("DB", "LD", "LO"),
        "species": ("Deer",),
        "sex_type": ("Buck",),
    },
    {
        "family": "ELK_ANTLERLESS",
        "output": "2026 elk antlerless all reviewed res-nr-total.csv",
        "sources": ["2026 ELK ANTLERLESS DRAW.xlsx"],
        "code_prefixes": ("EA",),
        "species": ("Elk",),
        "sex_type": ("Antlerless",),
    },
    {
        "family": "ELK_BULL",
        "output": "2026 elk bull all reviewed res-nr-total.csv",
        "sources": ["2026 ELK BULL ALL HUNTS.xlsx"],
        "code_prefixes": ("EB", "EL", "LO"),
        "species": ("Elk",),
        "sex_type": ("Bull",),
    },
    {
        "family": "MOOSE_BULL",
        "output": "2026 moose bull all reviewed res-nr-total.csv",
        "sources": ["2026 MOOSE BULL O.I.L.xlsx"],
        "code_prefixes": ("MB",),
        "species": ("Moose",),
        "sex_type": ("Bull",),
    },
    {
        "family": "MOOSE_ANTLERLESS",
        "output": "2026 moose antlerless reviewed res-nr-total.csv",
        "source_label": "USER_PASTED_DWR_HUNT_PLANNER_MOOSE_ANTLERLESS_MA_2026",
        "inline_tsv": """
Morgan-South Rich	MA1005	Antlerless	Moose	Any Legal Weapon	Limited Entry	Sep 20 2025 - Oct 19 2025	Res: 6
NonRes: 1
Morgan-South Rich	MA1007	Antlerless	Moose	Archery	Limited Entry	Aug 01 2025 - Sep 19 2025	Res: 3
NonRes: 0
Wasatch Mtns, Salt Lake-Timpanogos	MA1008	Antlerless	Moose	Any Legal Weapon	Limited Entry	Sep 20 2025 - Oct 19 2025	Res: 4
NonRes: 1
Causey Spring CWMU	MA1009	Antlerless	Moose	Any Legal Weapon	CWMU	Contact Operator for 2025 season dates	Res: 1
NonRes: 0
Deseret CWMU	MA1010	Antlerless	Moose	Any Legal Weapon	CWMU	Contact Operator for 2025 season dates	Res: 2
NonRes: 0
        """,
        "code_prefixes": ("MA",),
        "species": ("Moose",),
        "sex_type": ("Antlerless",),
    },
    {
        "family": "MOUNTAIN_GOAT_HUNTERS_CHOICE",
        "output": "2026 mountain goat reviewed res-nr-total.csv",
        "sources": ["2026 MOUNTAIN GOAT HUNTER CHOICE O.I.L.xlsx"],
        "code_prefixes": ("GO",),
        "species": ("Mountain Goat",),
        "sex_type": ("Hunters Choice",),
    },
    {
        "family": "PRONGHORN_BUCK",
        "output": "2026 pronghorn buck all reviewed res-nr-total.csv",
        "source_label": "USER_PASTED_DWR_HUNT_PLANNER_PRONGHORN_BUCK_PB_LP_2026",
        "inline_tsv": """
Beaver - Private Land Only	LP5025	Buck	Pronghorn	Any Legal Weapon	Limited Entry - Private Land Only	Sept 13 2025 - Sept 21 2025 - Private Land Only
Box Elder, West - Private Land Only	LP5031	Buck	Pronghorn	Any Legal Weapon	Limited Entry - Private Land Only	Sept 13 2025 - Sept 21 2025 - Private Land Only
Fillmore, Oak Creek South - Private Land Only	LP5033	Buck	Pronghorn	Any Legal Weapon	Limited Entry - Private Land Only	Sept 13 2025 - Sept 21 2025 - Private Land Only
San Rafael, North - Private Land Only	LP5046	Buck	Pronghorn	Any Legal Weapon	Limited Entry - Private Land Only	Sept 13 2025 - Sept 21 2025 - Private Land Only
Southwest Desert - Private Land Only	LP5049	Buck	Pronghorn	Any Legal Weapon	Limited Entry - Private Land Only	Sept 13 2025 - Sept 21 2025 - Private Land Only
West Desert, Rush Valley - Private Land Only	LP5051	Buck	Pronghorn	Any Legal Weapon	Limited Entry - Private Land Only	Sept 13 2025 - Sept 21 2025 - Private Land Only
Pronghorn - Statewide Permit	PB1000	Buck	Pronghorn	Any Legal Weapon	Statewide	Archery: Aug 15 - Aug 31, 2026 | ALW: Sept 1 - Nov 15, 2026
Beaver	PB5000	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 11
NonRes: 1
Book Cliffs, Bitter Creek	PB5001	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 13
NonRes: 1
Book Cliffs, South	PB5002	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 8
NonRes: 1
Box Elder, Promontory	PB5003	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 4
NonRes: 0
Box Elder, Puddle Valley	PB5004	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 3
NonRes: 1
Box Elder, Snowville	PB5005	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 9
NonRes: 1
Box Elder, West	PB5006	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 10
NonRes: 1
Fillmore, Oak Creek South	PB5008	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 15
NonRes: 2
La Sal, Potash/South Cisco	PB5009	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 6
NonRes: 0
Nine Mile, Anthro-Myton Bench	PB5011	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 30
NonRes: 3
North Slope, Three Corners/West Daggett	PB5012	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 10
NonRes: 1
Pine Valley	PB5013	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 7
NonRes: 1
San Rafael, North	PB5015	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 21
NonRes: 2
Diamond Mtn/Bonanza	PB5016	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 17
NonRes: 2
Vernal	PB5017	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 18
NonRes: 2
Southwest Desert	PB5018	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 31
NonRes: 4
West Desert, Riverbed	PB5019	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 18
NonRes: 2
West Desert, Rush Valley	PB5020	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 8
NonRes: 1
West Desert, Snake Valley	PB5021	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 23
NonRes: 3
Southwest Desert	PB5024	Buck	Pronghorn	Muzzleloader	Limited Entry	Sept 23 2026 - Oct 01 2026	Res: 31
NonRes: 4
Beaver	PB5025	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 43
NonRes: 5
Book Cliffs, Bitter Creek	PB5026	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 50
NonRes: 6
Book Cliffs, South	PB5027	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 34
NonRes: 4
Box Elder, Promontory	PB5028	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 14
NonRes: 2
Box Elder, Puddle Valley	PB5029	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 14
NonRes: 2
Box Elder, Snowville	PB5030	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 28
NonRes: 3
Box Elder, West	PB5031	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 30
NonRes: 3
Fillmore, Oak Creek South	PB5033	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 46
NonRes: 5
Kaiparowits	PB5034	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 2
NonRes: 0
La Sal, Potash/South Cisco	PB5035	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 22
NonRes: 2
Nine Mile, Anthro-Myton Bench	PB5037	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 89
NonRes: 10
Nine Mile, Range Creek	PB5038	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 13
NonRes: 1
North Slope, Summit	PB5039	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 15
NonRes: 2
North Slope, Three Corners/West Daggett	PB5040	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 29
NonRes: 3
Panguitch Lake/Zion, North	PB5041	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 21
NonRes: 2
Pine Valley	PB5042	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 21
NonRes: 2
San Juan, Hatch Point	PB5044	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 3
NonRes: 0
San Rafael, Desert	PB5045	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 7
NonRes: 1
San Rafael, North	PB5046	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 62
NonRes: 7
Diamond Mtn/Bonanza	PB5047	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 50
NonRes: 6
Vernal	PB5048	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 54
NonRes: 6
Southwest Desert	PB5049	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 94
NonRes: 11
West Desert, Riverbed	PB5050	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 73
NonRes: 8
West Desert, Rush Valley	PB5051	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 32
NonRes: 4
West Desert, Snake Valley	PB5052	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 95
NonRes: 11
Nine Mile, Range Creek	PB5053	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 4
NonRes: 0
Panguitch Lake/Zion, North	PB5054	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 7
NonRes: 1
San Rafael, Desert	PB5055	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 2
NonRes: 0
San Rafael, North	PB5056	Buck	Pronghorn	Muzzleloader	Limited Entry	Sept 23 2026 - Oct 01 2026	Res: 21
NonRes: 2
Nine Mile, Anthro-Myton Bench	PB5059	Buck	Pronghorn	Muzzleloader	Limited Entry	Sept 23 2026 - Oct 01 2026	Res: 30
NonRes: 3
Diamond Mtn/Bonanza	PB5060	Buck	Pronghorn	Muzzleloader	Limited Entry	Sept 23 2026 - Oct 01 2026	Res: 17
NonRes: 2
Panguitch Lake/Zion, North	PB5061	Buck	Pronghorn	Muzzleloader	Limited Entry	Sept 23 2026 - Oct 01 2026	Res: 7
NonRes: 1
Pine Valley	PB5062	Buck	Pronghorn	Muzzleloader	Limited Entry	Sept 23 2026 - Oct 01 2026	Res: 7
NonRes: 1
Fillmore, Oak Creek South	PB5065	Buck	Pronghorn	Muzzleloader	Limited Entry	Sept 23 2026 - Oct 01 2026	Res: 15
NonRes: 2
North Slope, Three Corners/West Daggett	PB5066	Buck	Pronghorn	Muzzleloader	Limited Entry	Sept 23 2026 - Oct 01 2026	Res: 10
NonRes: 1
Vernal	PB5072	Buck	Pronghorn	Muzzleloader	Limited Entry	Sept 23 2026 - Oct 01 2026	Res: 18
NonRes: 2
Cache, Rich	PB5073	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 6
NonRes: 0
Cache, Rich	PB5074	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 2
NonRes: 0
Cache, Rich	PB5075	Buck	Pronghorn	Muzzleloader	Limited Entry	Sept 23 2026 - Oct 01 2026	Res: 2
NonRes: 0
Parker Mtn	PB5076	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 135
NonRes: 15
Parker Mtn	PB5077	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 45
NonRes: 5
Parker Mtn	PB5078	Buck	Pronghorn	Muzzleloader	Limited Entry	Sept 23 2026 - Oct 01 2026	Res: 45
NonRes: 5
Box Elder, West	PB5079	Buck	Pronghorn	Muzzleloader	Limited Entry	Sept 23 2026 - Oct 01 2026	Res: 10
NonRes: 1
Box Elder, Snowville	PB5080	Buck	Pronghorn	Muzzleloader	Limited Entry	Sept 23 2026 - Oct 01 2026	Res: 9
NonRes: 1
Antelope Creek CWMU	PB5300	Buck	Pronghorn	Any Legal Weapon	CWMU	Contact operator for 2026 season dates	Res: 4
NonRes: 0
Deseret CWMU	PB5302	Buck	Pronghorn	Any Legal Weapon	CWMU	Contact operator for 2026 season dates	Res: 29
NonRes: 0
Park Valley CWMU	PB5305	Buck	Pronghorn	Any Legal Weapon	CWMU	Contact operator for 2026 season dates	Res: 1
NonRes: 0
Rabbit Creek CWMU	PB5306	Buck	Pronghorn	Any Legal Weapon	CWMU	Contact operator for 2026 season dates	Res: 1
NonRes: 0
Skull Valley South CWMU	PB5309	Buck	Pronghorn	Any Legal Weapon	CWMU	Contact operator for 2026 season dates	Res: 2
NonRes: 0
The Rose of Snowville CWMU	PB5311	Buck	Pronghorn	Any Legal Weapon	CWMU	Contact operator for 2026 season dates	Res: 3
NonRes: 0
TL Bar Ranch CWMU	PB5312	Buck	Pronghorn	Any Legal Weapon	CWMU	Contact operator for 2026 season dates	Res: 1
NonRes: 0
Westlake CWMU	PB5314	Buck	Pronghorn	Any Legal Weapon	CWMU	Contact operator for 2026 season dates	Res: 2
NonRes: 0
Zane CWMU	PB5315	Buck	Pronghorn	Any Legal Weapon	CWMU	Contact operator for 2026 season dates	Res: 2
NonRes: 0
Allen Ranch CWMU	PB5325	Buck	Pronghorn	Any Legal Weapon	CWMU	Contact operator for 2026 season dates	Res: 2
NonRes: 0
Cedar Springs CWMU	PB5326	Buck	Pronghorn	Any Legal Weapon	CWMU	Contact operator for 2026 season dates	Res: 1
NonRes: 0
Cottonwood Ridge CWMU	PB5327	Buck	Pronghorn	Any Legal Weapon	CWMU	Contact operator for 2026 season dates	Res: 4
NonRes: 0
Pahvant Ensign CWMU	PB5328	Buck	Pronghorn	Any Legal Weapon	CWMU	Contact operator for 2026 season dates	Res: 3
NonRes: 0
Snowville Flat CWMU	PB5329	Buck	Pronghorn	Any Legal Weapon	CWMU	Contact operator for 2026 season dates	Res: 2
NonRes: 0
George Creek CWMU	PB5330	Buck	Pronghorn	Any Legal Weapon	CWMU	Contact operator for 2026 season dates	Res: 1
NonRes: 0
Mt Dutton/Paunsaugunt	PB5331	Buck	Pronghorn	Any Legal Weapon	Limited Entry	Sept 12 2026 - Sept 20 2026	Res: 13
NonRes: 2
Mt Dutton/Paunsaugunt	PB5332	Buck	Pronghorn	Archery	Limited Entry	Aug 15 2026 - Sept 11 2026	Res: 4
NonRes: 1
Heist CWMU	PB5337	Buck	Pronghorn	Any Legal Weapon	CWMU	Contact operator for 2026 season dates	Res: 2
NonRes: 0
Castle Valley Outdoors CWMU	PB5338	Buck	Pronghorn	Any Legal Weapon	CWMU	Contact operator for 2026 season dates	Res: 1
NonRes: 0
Iron Spring CWMU	PB5339	Buck	Pronghorn	Any Legal Weapon	CWMU	Contact operator for 2026 season dates	Res: 1
NonRes: 0
Salt Wells CWMU	PB5342	Buck	Pronghorn	Any Legal Weapon	CWMU	Contact operator for 2026 season dates	Res: 1
NonRes: 0
Green River Flat CWMU	PB5345	Buck	Pronghorn	Any Legal Weapon	CWMU	Contact operator for 2026 season dates	Res: 2
NonRes: 0
Junction Valley CWMU	PB5346	Buck	Pronghorn	Any Legal Weapon	CWMU	Contact operator for 2026 season dates	Res: 1
NonRes: 0
        """,
        "code_prefixes": ("PB", "LP"),
        "species": ("Pronghorn",),
        "sex_type": ("Buck",),
    },
    {
        "family": "PRONGHORN_DOE",
        "output": "2026 pronghorn doe all reviewed res-nr-total.csv",
        "source_label": "USER_PASTED_DWR_HUNT_PLANNER_PRONGHORN_DOE_PD_2026",
        "inline_rows": [
            {"hunt_name": "Box Elder, Snowville", "hunt_code": "PD1000", "sex_type": "Doe", "species": "Pronghorn", "weapon": "Any Legal Weapon", "hunt_type": "General Season", "season": "Nov 08 2025 - Dec 07 2025", "permits_2026_res": "72", "permits_2026_nr": "8"},
            {"hunt_name": "Box Elder, Promontory", "hunt_code": "PD1012", "sex_type": "Doe", "species": "Pronghorn", "weapon": "Any Legal Weapon", "hunt_type": "General Season", "season": "Oct 04 2025 - Oct 26 2025", "permits_2026_res": "22", "permits_2026_nr": "3"},
            {"hunt_name": "Box Elder, West", "hunt_code": "PD1017", "sex_type": "Doe", "species": "Pronghorn", "weapon": "Any Legal Weapon", "hunt_type": "General Season", "season": "Oct 04 2025 - Oct 26 2025", "permits_2026_res": "49", "permits_2026_nr": "6"},
            {"hunt_name": "Westlake CWMU", "hunt_code": "PD1031", "sex_type": "Doe", "species": "Pronghorn", "weapon": "Any Legal Weapon", "hunt_type": "CWMU", "season": "Contact operator for 2025 season dates", "permits_2026_res": "9", "permits_2026_nr": "0"},
            {"hunt_name": "San Rafael, Poison Springs Bench", "hunt_code": "PD1033", "sex_type": "Doe", "species": "Pronghorn", "weapon": "Any Legal Weapon", "hunt_type": "General Season", "season": "Oct 04 2025 - Oct 26 2025", "permits_2026_res": "13", "permits_2026_nr": "2"},
            {"hunt_name": "Fillmore, Oak Creek South", "hunt_code": "PD1034", "sex_type": "Doe", "species": "Pronghorn", "weapon": "Any Legal Weapon", "hunt_type": "General Season", "season": "Aug 01 2025 - Aug 13 2025", "permits_2026_res": "27", "permits_2026_nr": "3"},
            {"hunt_name": "Fillmore, Oak Creek South", "hunt_code": "PD1035", "sex_type": "Doe", "species": "Pronghorn", "weapon": "Any Legal Weapon", "hunt_type": "General Season", "season": "Nov 01 2025 - Nov 29 2025", "permits_2026_res": "27", "permits_2026_nr": "3"},
            {"hunt_name": "Nine Mile, Anthro-Pleasant Valley", "hunt_code": "PD1043", "sex_type": "Doe", "species": "Pronghorn", "weapon": "Any Legal Weapon", "hunt_type": "General Season", "season": "Aug 01 2025 - Aug 16 2025", "permits_2026_res": "9", "permits_2026_nr": "1"},
            {"hunt_name": "Vernal-Ouray Valley", "hunt_code": "PD1044", "sex_type": "Doe", "species": "Pronghorn", "weapon": "Any Legal Weapon", "hunt_type": "General Season", "season": "Oct 04 2025 - Oct 16 2025", "permits_2026_res": "9", "permits_2026_nr": "1"},
            {"hunt_name": "Antelope Creek CWMU", "hunt_code": "PD1047", "sex_type": "Doe", "species": "Pronghorn", "weapon": "Any Legal Weapon", "hunt_type": "CWMU", "season": "Contact operator for 2025 season dates", "permits_2026_res": "6", "permits_2026_nr": "0"},
            {"hunt_name": "The Rose of Snowville CWMU", "hunt_code": "PD1048", "sex_type": "Doe", "species": "Pronghorn", "weapon": "Any Legal Weapon", "hunt_type": "CWMU", "season": "Contact operator for 2025 season dates", "permits_2026_res": "2", "permits_2026_nr": "0"},
            {"hunt_name": "Zane CWMU", "hunt_code": "PD1049", "sex_type": "Doe", "species": "Pronghorn", "weapon": "Any Legal Weapon", "hunt_type": "CWMU", "season": "Contact operator for 2025 season dates", "permits_2026_res": "6", "permits_2026_nr": "0"},
            {"hunt_name": "Cottonwood Ridge CWMU", "hunt_code": "PD1050", "sex_type": "Doe", "species": "Pronghorn", "weapon": "Any Legal Weapon", "hunt_type": "CWMU", "season": "Contact operator for 2025 season dates", "permits_2026_res": "6", "permits_2026_nr": "0"},
            {"hunt_name": "Pahvant Ensign CWMU", "hunt_code": "PD1051", "sex_type": "Doe", "species": "Pronghorn", "weapon": "Any Legal Weapon", "hunt_type": "CWMU", "season": "Contact operator for 2025 season dates", "permits_2026_res": "6", "permits_2026_nr": "0"},
            {"hunt_name": "Heist CWMU", "hunt_code": "PD1052", "sex_type": "Doe", "species": "Pronghorn", "weapon": "Any Legal Weapon", "hunt_type": "CWMU", "season": "Contact operator for 2025 season dates", "permits_2026_res": "6", "permits_2026_nr": "0"},
            {"hunt_name": "Green River Flat CWMU", "hunt_code": "PD1053", "sex_type": "Doe", "species": "Pronghorn", "weapon": "Any Legal Weapon", "hunt_type": "CWMU", "season": "Contact operator for 2025 season dates", "permits_2026_res": "6", "permits_2026_nr": "0"},
            {"hunt_name": "Parker Mtn", "hunt_code": "PD1054", "sex_type": "Doe", "species": "Pronghorn", "weapon": "Archery", "hunt_type": "General Season", "season": "Aug 30 2025 - Sep 12 2025", "permits_2026_res": "27", "permits_2026_nr": "3"},
            {"hunt_name": "Parker Mtn", "hunt_code": "PD1055", "sex_type": "Doe", "species": "Pronghorn", "weapon": "Muzzleloader", "hunt_type": "General Season", "season": "Oct 29 2025 - Nov 08 2025", "permits_2026_res": "31", "permits_2026_nr": "4"},
            {"hunt_name": "Parker Mtn", "hunt_code": "PD1056", "sex_type": "Doe", "species": "Pronghorn", "weapon": "Any Legal Weapon", "hunt_type": "General Season", "season": "Dec 13 2025 - Dec 21 2025", "permits_2026_res": "36", "permits_2026_nr": "4"},
            {"hunt_name": "Parker Mtn, Highlands", "hunt_code": "PD1057", "sex_type": "Doe", "species": "Pronghorn", "weapon": "Any Legal Weapon", "hunt_type": "General Season", "season": "Oct 08 2025 - Oct 16 2025", "permits_2026_res": "31", "permits_2026_nr": "4"},
            {"hunt_name": "Parker Mtn, Highlands", "hunt_code": "PD1058", "sex_type": "Doe", "species": "Pronghorn", "weapon": "Any Legal Weapon", "hunt_type": "General Season", "season": "Nov 22 2025 - Nov 30 2025", "permits_2026_res": "36", "permits_2026_nr": "4"},
            {"hunt_name": "Parker Mtn, Plains", "hunt_code": "PD1059", "sex_type": "Doe", "species": "Pronghorn", "weapon": "Any Legal Weapon", "hunt_type": "General Season", "season": "Oct 08 2025 - Oct 16 2025", "permits_2026_res": "27", "permits_2026_nr": "3"},
            {"hunt_name": "Parker Mtn, Plains", "hunt_code": "PD1060", "sex_type": "Doe", "species": "Pronghorn", "weapon": "Any Legal Weapon", "hunt_type": "General Season", "season": "Nov 22 2025 - Nov 30 2025", "permits_2026_res": "36", "permits_2026_nr": "4"},
        ],
        "code_prefixes": ("PD",),
        "species": ("Pronghorn",),
        "sex_type": ("Doe",),
    },
    {
        "family": "TURKEY_BEARDED",
        "output": "2026 turkey bearded all reviewed total.csv",
        "sources": ["2026 TURKEY BEARDED DRAW.xlsx"],
        "code_prefixes": ("TK",),
        "species": ("Turkey",),
        "sex_type": ("Bearded",),
    },
    {
        "family": "TURKEY_EITHER_SEX",
        "output": "2026 turkey either sex all reviewed total.csv",
        "sources": ["2026 TURKEY EITHER SEX DRAW.xlsx"],
        "code_prefixes": ("TK",),
        "species": ("Turkey",),
        "sex_type": ("Either Sex",),
    },
    {
        "family": "ROCKY_MOUNTAIN_BIGHORN_RAM",
        "output": "2026 rocky mountain bighorn reviewed res-nr-total.csv",
        "source_label": "USER_PASTED_DWR_HUNT_PLANNER_ROCKY_MOUNTAIN_BIGHORN_RS_2026",
        "inline_rows": [
            {
                "hunt_name": "Rocky Mountain Bighorn Sheep - Statewide Permit",
                "hunt_code": "RS0001",
                "sex_type": "Male Only",
                "species": "Rocky Mountain Bighorn Sheep",
                "weapon": "Any Legal Weapon",
                "hunt_type": "Statewide",
                "season": "Sept 1 - Dec 31, 2026",
            },
            {
                "hunt_name": "Antelope Island Conservation",
                "hunt_code": "RS1000",
                "sex_type": "Male Only",
                "species": "Rocky Mountain Bighorn Sheep",
                "weapon": "Any Legal Weapon",
                "hunt_type": "Once-in-a-lifetime",
                "season": "Nov 9 2026 - Nov 16 2026",
            },
            {
                "hunt_name": "Book Cliffs, South",
                "hunt_code": "RS1001",
                "sex_type": "Male Only",
                "species": "Rocky Mountain Bighorn Sheep",
                "weapon": "Any Legal Weapon",
                "hunt_type": "Conservation",
                "season": "Nov 1 - Dec 31, 2025",
            },
            {
                "hunt_name": "Box Elder, Newfoundland Mtn",
                "hunt_code": "RS1003",
                "sex_type": "Male Only",
                "species": "Rocky Mountain Bighorn Sheep",
                "weapon": "Any Legal Weapon",
                "hunt_type": "Conservation",
                "season": "Oct 25 - Dec 31, 2025",
            },
            {
                "hunt_name": "Nine Mile, Gray Canyon",
                "hunt_code": "RS1006",
                "sex_type": "Male Only",
                "species": "Rocky Mountain Bighorn Sheep",
                "weapon": "Any Legal Weapon",
                "hunt_type": "Conservation",
                "season": "Nov 1 - Dec 31, 2025",
            },
            {
                "hunt_name": "Antelope Island",
                "hunt_code": "RS6700",
                "sex_type": "Male Only",
                "species": "Rocky Mountain Bighorn Sheep",
                "weapon": "Any Legal Weapon",
                "hunt_type": "Once-in-a-lifetime",
                "season": "Nov 11, 2026 - Nov 18, 2026",
                "permits_2026_res": "2",
                "permits_2026_nr": "0",
            },
            {
                "hunt_name": "Book Cliffs, South",
                "hunt_code": "RS6701",
                "sex_type": "Male Only",
                "species": "Rocky Mountain Bighorn Sheep",
                "weapon": "Any Legal Weapon",
                "hunt_type": "Once-in-a-lifetime",
                "season": "Oct 31 2026 - Nov 30 2026",
                "permits_2026_res": "5",
                "permits_2026_nr": "1",
            },
            {
                "hunt_name": "Box Elder, Newfoundland Mtn",
                "hunt_code": "RS6703",
                "sex_type": "Male Only",
                "species": "Rocky Mountain Bighorn Sheep",
                "weapon": "Any Legal Weapon",
                "hunt_type": "Once-in-a-lifetime",
                "season": "Oct 03 2026 - Oct 23 2026",
                "permits_2026_res": "4",
                "permits_2026_nr": "0",
            },
            {
                "hunt_name": "Box Elder, Newfoundland Mtn",
                "hunt_code": "RS6704",
                "sex_type": "Male Only",
                "species": "Rocky Mountain Bighorn Sheep",
                "weapon": "Any Legal Weapon",
                "hunt_type": "Once-in-a-lifetime",
                "season": "Oct 24 2026 - Nov 13 2026",
                "permits_2026_res": "3",
                "permits_2026_nr": "1",
            },
            {
                "hunt_name": "North Slope, Three Corners",
                "hunt_code": "RS6708",
                "sex_type": "Male Only",
                "species": "Rocky Mountain Bighorn Sheep",
                "weapon": "Any Legal Weapon",
                "hunt_type": "Once-in-a-lifetime",
                "season": "Sept 14 2026 - Nov 30 2026",
                "permits_2026_res": "2",
                "permits_2026_nr": "0",
            },
            {
                "hunt_name": "North Slope, Summit/West Daggett",
                "hunt_code": "RS6709",
                "sex_type": "Male Only",
                "species": "Rocky Mountain Bighorn Sheep",
                "weapon": "Any Legal Weapon",
                "hunt_type": "Once-in-a-lifetime",
                "season": "Sept 14 2026 - Nov 30 2026",
                "permits_2026_res": "3",
                "permits_2026_nr": "0",
            },
            {
                "hunt_name": "Nine Mile, Gray Canyon",
                "hunt_code": "RS6712",
                "sex_type": "Male Only",
                "species": "Rocky Mountain Bighorn Sheep",
                "weapon": "Any Legal Weapon",
                "hunt_type": "Once-in-a-lifetime",
                "season": "Oct 31 2026 - Nov 30 2026",
                "permits_2026_res": "4",
                "permits_2026_nr": "2",
            },
            {
                "hunt_name": "Nine Mile, Jack Creek",
                "hunt_code": "RS6713",
                "sex_type": "Male Only",
                "species": "Rocky Mountain Bighorn Sheep",
                "weapon": "Any Legal Weapon",
                "hunt_type": "Once-in-a-lifetime",
                "season": "Oct 31 2026 - Nov 30 2026",
                "permits_2026_res": "4",
                "permits_2026_nr": "0",
            },
            {
                "hunt_name": "Fillmore, Oak Creek",
                "hunt_code": "RS6720",
                "sex_type": "Male Only",
                "species": "Rocky Mountain Bighorn Sheep",
                "weapon": "Any Legal Weapon",
                "hunt_type": "Once-in-a-lifetime",
                "season": "Oct 03 2026 - Oct 23 2026",
                "permits_2026_res": "4",
                "permits_2026_nr": "0",
            },
            {
                "hunt_name": "Oquirrh-Stansbury, West",
                "hunt_code": "RS6721",
                "sex_type": "Male Only",
                "species": "Rocky Mountain Bighorn Sheep",
                "weapon": "Any Legal Weapon",
                "hunt_type": "Once-in-a-lifetime",
                "season": "Oct 31 2026 - Nov 30 2026",
                "permits_2026_res": "3",
                "permits_2026_nr": "0",
            },
            {
                "hunt_name": "Box Elder, Newfoundland Mtn",
                "hunt_code": "RS6722",
                "sex_type": "Male Only",
                "species": "Rocky Mountain Bighorn Sheep",
                "weapon": "Archery",
                "hunt_type": "Once-in-a-lifetime",
                "season": "Nov 14 2026 - Dec 06 2026",
                "permits_2026_res": "2",
                "permits_2026_nr": "0",
            },
            {
                "hunt_name": "Wasatch Mtns, West",
                "hunt_code": "RS6724",
                "sex_type": "Male Only",
                "species": "Rocky Mountain Bighorn Sheep",
                "weapon": "Any Legal Weapon",
                "hunt_type": "Once-in-a-lifetime",
                "season": "Oct 31 2026 - Nov 30 2026",
                "permits_2026_res": "3",
                "permits_2026_nr": "0",
            },
            {
                "hunt_name": "Nebo",
                "hunt_code": "RS6725",
                "sex_type": "Male Only",
                "species": "Rocky Mountain Bighorn Sheep",
                "weapon": "Any Legal Weapon",
                "hunt_type": "Once-in-a-lifetime",
                "season": "Oct 31 2026 - Nov 30 2026",
                "permits_2026_res": "2",
                "permits_2026_nr": "0",
            },
            {
                "hunt_name": "Fillmore, Oak Creek",
                "hunt_code": "RS6726",
                "sex_type": "Male Only",
                "species": "Rocky Mountain Bighorn Sheep",
                "weapon": "Any Legal Weapon",
                "hunt_type": "Once-in-a-lifetime",
                "season": "Oct 24 2026 - Nov 13 2026",
                "permits_2026_res": "3",
                "permits_2026_nr": "1",
            },
            {
                "hunt_name": "Fillmore, Oak Creek",
                "hunt_code": "RS6727",
                "sex_type": "Male Only",
                "species": "Rocky Mountain Bighorn Sheep",
                "weapon": "Archery",
                "hunt_type": "Once-in-a-lifetime",
                "season": "Nov 14 2026 - Dec 06 2026",
                "permits_2026_res": "2",
                "permits_2026_nr": "0",
            },
        ],
    },
    {
        "family": "ROCKY_MOUNTAIN_BIGHORN_EWE",
        "output": "2026 rocky mountain bighorn ewe reviewed res-nr-total.csv",
        "sources": ["2026 ROCKY MOUNTAIN BIGHORN SHEEP EWE O.I.L.xlsx"],
        "code_prefixes": ("RE",),
        "species": ("Rocky Mountain Bighorn Sheep",),
        "sex_type": ("Ewe",),
    },
    {
        "family": "DESERT_BIGHORN_RAM",
        "output": "2026 desert bighorn reviewed res-nr-total.csv",
        "sources": ["2026 DESERT BIGHORN SHEEP RAM O.I.L.xlsx"],
        "code_prefixes": ("DS",),
        "species": ("Desert Bighorn Sheep",),
        "sex_type": ("Male Only",),
        "blank_permit_codes": ("DS1000", "DS1002", "DS1003", "DS1004", "DS1006", "DS1007", "DS6605"),
    },
]


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def norm(value: object) -> str:
    return re.sub(r"\s+", " ", text(value)).casefold()


def code(value: object) -> str:
    return re.sub(r"[^A-Z0-9]", "", text(value).upper())


def numeric(value: object) -> str:
    raw = text(value)
    if not raw:
        return ""
    match = re.search(r"-?\d+(?:\.\d+)?", raw)
    if not match:
        return ""
    number = float(match.group(0))
    return str(int(number)) if number.is_integer() else str(number)


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [{k: text(v) for k, v in row.items()} for row in csv.DictReader(handle)]


def read_database_map() -> dict[str, dict[str, str]]:
    rows = read_csv_rows(DATABASE_CSV)
    return {code(row.get("hunt_code")): row for row in rows if code(row.get("hunt_code"))}


def read_xlsx_rows(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [text(value) for value in rows[0]]
    output: list[dict[str, object]] = []
    for values in rows[1:]:
        row = dict(zip(headers, values))
        if text(row.get("hunt_code")):
            output.append(row)
    return output


def parse_permit_value(value: str) -> tuple[str, str] | None:
    match = re.match(r"^(Res|NonRes|Total):\s*(\d+)\s*$", value.strip(), flags=re.I)
    if not match:
        return None
    key = match.group(1).lower()
    if key == "nonres":
        return "permits_2026_nr", match.group(2)
    if key == "res":
        return "permits_2026_res", match.group(2)
    return "permits_2026_total", match.group(2)


def parse_inline_tsv_rows(raw: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for line in raw.splitlines():
        line = line.strip()
        if not line:
            continue

        single_value = parse_permit_value(line)
        if single_value and rows:
            field, value = single_value
            rows[-1][field] = value
            continue

        parts = [part.strip() for part in line.split("\t")]
        if len(parts) < 7:
            raise ValueError(f"Inline permit row has fewer than 7 fields: {line}")

        row = {
            "hunt_name": parts[0],
            "hunt_code": parts[1],
            "sex_type": parts[2],
            "species": parts[3],
            "weapon": parts[4],
            "hunt_type": parts[5],
            "season": parts[6],
        }
        for value in parts[7:]:
            parsed = parse_permit_value(value)
            if parsed:
                field, amount = parsed
                row[field] = amount
        rows.append(row)
    return rows


def permit_status(res: str, nr: str, total: str) -> str:
    if res and nr and total:
        return "FULL_SPLIT"
    if total and not res and not nr:
        return "TOTAL_ONLY"
    if not total and not res and not nr:
        return "NO_PUBLISHED_NUMERIC_PERMIT"
    return "PARTIAL_PERMIT_FIELDS_REVIEW"


def row_allowed(row: dict[str, object], config: dict[str, object]) -> bool:
    hunt_code = code(row.get("hunt_code"))
    prefixes = tuple(config.get("code_prefixes", ()))
    if prefixes and not hunt_code.startswith(prefixes):
        return False
    species = tuple(norm(v) for v in config.get("species", ()))
    if species and norm(row.get("species")) not in species:
        return False
    sex_types = tuple(norm(v) for v in config.get("sex_type", ()))
    if sex_types and norm(row.get("sex_type")) not in sex_types:
        return False
    return True


def canonicalize_row(
    row: dict[str, object],
    source_file: str,
    source_row_number: int,
    database: dict[str, dict[str, str]],
) -> dict[str, str]:
    hunt_code = code(row.get("hunt_code"))
    database_row = database.get(hunt_code, {})
    boundary_id = text(database_row.get("boundary_id") or database_row.get("BOUNDARY_ID"))

    res = numeric(row.get("permits_2026_res"))
    nr = numeric(row.get("permits_2026_nr"))
    total = numeric(row.get("permits_2026_total"))
    if not total and res and nr:
        total = str(int(res) + int(nr))

    return {
        "hunt_name": text(row.get("hunt_name")),
        "hunt_code": hunt_code,
        "boundary_id": boundary_id,
        "hunt_code_mapping_status": "REVIEWED_CURRENT_HUNT_CODE",
        "boundary_id_mapping_status": "DATABASE_BOUNDARY_ID" if boundary_id else "MISSING_BOUNDARY_ID_REVIEW",
        "candidate_hunt_code": hunt_code,
        "candidate_boundary_id": boundary_id,
        "sex_type": text(row.get("sex_type")),
        "species": text(row.get("species")),
        "weapon": text(row.get("weapon")),
        "hunt_type": text(row.get("hunt_type")),
        "season": text(row.get("season")),
        "permits_2026_res": res,
        "permits_2026_nr": nr,
        "permits_2026_total": total,
        "permit_count_status": permit_status(res, nr, total),
        "source_file": source_file,
        "source_row_number": str(source_row_number),
        "_source_file": source_file,
        "_source_row_number": str(source_row_number),
    }


def sort_key(row: dict[str, str]) -> tuple[str, int, str]:
    hunt_code = row["hunt_code"]
    match = re.match(r"([A-Z]+)(\d+)", hunt_code)
    if not match:
        return (hunt_code, 0, hunt_code)
    return (match.group(1), int(match.group(2)), hunt_code)


def write_truth_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CANONICAL_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in CANONICAL_COLUMNS})


def write_master_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=MASTER_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in MASTER_COLUMNS})


def catalog_superseded_fragments() -> list[dict[str, str]]:
    manifest: list[dict[str, str]] = []
    for file_name in SUPERSEDED_FRAGMENT_FILES:
        active_path = PERMIT_DIR / file_name
        status = "ACTIVE_SUPERSEDED_FRAGMENT" if active_path.exists() else "MISSING"
        if active_path.exists():
            status = "ACTIVE_SUPERSEDED_FRAGMENT_DO_NOT_USE_FOR_CANONICAL_PERMITS"

        manifest.append(
            {
                "file_name": file_name,
                "active_path": str(active_path),
                "status": status,
                "reason": "Superseded by canonical reviewed family outputs and 2026 reviewed permit truth master.csv",
            }
        )
    return manifest


def write_fragment_manifest(rows: list[dict[str, str]]) -> None:
    fieldnames = ["file_name", "active_path", "status", "reason"]
    with SUPERSEDED_FRAGMENT_MANIFEST.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def validate_family(rows: list[dict[str, str]]) -> dict[str, object]:
    codes = [row["hunt_code"] for row in rows]
    duplicates = sorted(code for code, count in Counter(codes).items() if count > 1)
    bad_totals = []
    for row in rows:
        res = numeric(row.get("permits_2026_res"))
        nr = numeric(row.get("permits_2026_nr"))
        total = numeric(row.get("permits_2026_total"))
        if res and nr and total and int(res) + int(nr) != int(total):
            bad_totals.append(row["hunt_code"])

    status_counts = Counter(row["permit_count_status"] for row in rows)
    hunt_type_counts = Counter(row["hunt_type"] for row in rows)
    weapon_counts = Counter(row["weapon"] for row in rows)
    source_counts = Counter(row["_source_file"] for row in rows)
    missing_source_rows = [
        row["hunt_code"] for row in rows if not row.get("source_file") or not row.get("source_row_number")
    ]
    return {
        "rows": len(rows),
        "duplicate_hunt_codes": duplicates,
        "bad_split_totals": bad_totals,
        "missing_source_rows": missing_source_rows,
        "permit_status_counts": dict(status_counts),
        "hunt_type_counts": dict(hunt_type_counts),
        "weapon_counts": dict(weapon_counts),
        "source_counts": dict(source_counts),
    }


def apply_config_overrides(row: dict[str, str], config: dict[str, object]) -> dict[str, str]:
    blank_codes = {code(value) for value in config.get("blank_permit_codes", ())}
    if row["hunt_code"] in blank_codes:
        row["permits_2026_res"] = ""
        row["permits_2026_nr"] = ""
        row["permits_2026_total"] = ""
        row["permit_count_status"] = permit_status("", "", "")

    if "statewide permit" in norm(row.get("hunt_name")):
        row["permits_2026_res"] = ""
        row["permits_2026_nr"] = ""
        row["permits_2026_total"] = "1"
        row["permit_count_status"] = permit_status("", "", "1")
    return row


def main() -> None:
    PERMIT_DIR.mkdir(parents=True, exist_ok=True)
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    database = read_database_map()

    audit_rows = []
    master_rows: list[dict[str, str]] = []
    audit_json = {
        "created_at": datetime.now(timezone.utc).isoformat(),
        "source": "pasted-aligned public hunt table workbooks plus reviewed inline pasted DWR Hunt Planner rows",
        "families": {},
    }

    for config in FAMILY_CONFIGS:
        family = str(config["family"])
        all_rows: list[dict[str, str]] = []
        if config.get("inline_rows"):
            source = str(config.get("source_label", "INLINE_REVIEWED_SOURCE_ROWS"))
            for index, row in enumerate(config["inline_rows"], start=1):
                if not row_allowed(row, config):
                    continue
                all_rows.append(apply_config_overrides(canonicalize_row(row, source, index, database), config))

        if config.get("inline_tsv"):
            source = str(config.get("source_label", "INLINE_REVIEWED_SOURCE_ROWS"))
            for index, row in enumerate(parse_inline_tsv_rows(str(config["inline_tsv"])), start=1):
                if not row_allowed(row, config):
                    continue
                all_rows.append(apply_config_overrides(canonicalize_row(row, source, index, database), config))

        for source in config.get("sources", []):
            source_path = PUBLIC_XLSX_DIR / source
            rows = read_xlsx_rows(source_path)
            for index, row in enumerate(rows, start=2):
                if not row_allowed(row, config):
                    continue
                all_rows.append(apply_config_overrides(canonicalize_row(row, source, index, database), config))

        all_rows.sort(key=sort_key)
        output_path = PERMIT_DIR / str(config["output"])
        write_truth_csv(output_path, all_rows)
        for row in all_rows:
            master_rows.append(
                {"reviewed_family": family, **{column: row.get(column, "") for column in CANONICAL_COLUMNS}}
            )

        validation = validate_family(all_rows)
        validation["output"] = str(output_path)
        audit_json["families"][family] = validation
        audit_rows.append(
            {
                "family": family,
                "output": str(output_path),
                "rows": validation["rows"],
                "full_split": validation["permit_status_counts"].get("FULL_SPLIT", 0),
                "total_only": validation["permit_status_counts"].get("TOTAL_ONLY", 0),
                "blank_no_published_numeric": validation["permit_status_counts"].get(
                    "NO_PUBLISHED_NUMERIC_PERMIT", 0
                ),
                "partial": validation["permit_status_counts"].get("PARTIAL_PERMIT_FIELDS_REVIEW", 0),
                "duplicate_hunt_codes": len(validation["duplicate_hunt_codes"]),
                "bad_split_totals": len(validation["bad_split_totals"]),
                "missing_source_rows": len(validation["missing_source_rows"]),
                "source_counts": json.dumps(validation["source_counts"], sort_keys=True),
            }
        )

    master_rows.sort(key=lambda row: (row["reviewed_family"], *sort_key(row)))
    write_master_csv(MASTER_REVIEWED_PERMIT_CSV, master_rows)
    fragment_manifest_rows = catalog_superseded_fragments()
    write_fragment_manifest(fragment_manifest_rows)

    reviewed_codes = {
        code(row.get("hunt_code"))
        for config in FAMILY_CONFIGS
        for row in (
            parse_inline_tsv_rows(str(config["inline_tsv"]))
            if config.get("inline_tsv")
            else config.get("inline_rows", [])
        )
        if code(row.get("hunt_code"))
    }
    for output in audit_json["families"].values():
        output_path = Path(str(output["output"]))
        if output_path.exists():
            reviewed_codes.update(code(row.get("hunt_code")) for row in read_csv_rows(output_path))

    excluded_codes = {code(item["hunt_code"]) for item in REVIEWED_DATABASE_EXCLUSIONS}
    database_codes = set(database)
    audit_json["database_reconciliation"] = {
        "database_hunt_code_count": len(database_codes),
        "reviewed_permit_hunt_code_count": len(reviewed_codes),
        "reviewed_database_exclusions": REVIEWED_DATABASE_EXCLUSIONS,
        "unresolved_database_hunt_codes_excluding_reviewed_exclusions": sorted(
            database_codes - reviewed_codes - excluded_codes
        ),
    }
    audit_json["canonical_outputs"] = {
        "master_reviewed_permit_csv": str(MASTER_REVIEWED_PERMIT_CSV),
        "master_rows": len(master_rows),
        "superseded_fragment_manifest": str(SUPERSEDED_FRAGMENT_MANIFEST),
        "superseded_fragment_status_counts": dict(Counter(row["status"] for row in fragment_manifest_rows)),
        "cleanup_mode": "NON_DESTRUCTIVE_MANIFEST_ONLY",
    }

    audit_csv_path = AUDIT_DIR / "reviewed_permit_truth_sources_2026_audit.csv"
    with audit_csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(audit_rows[0].keys()))
        writer.writeheader()
        writer.writerows(audit_rows)

    audit_json_path = AUDIT_DIR / "reviewed_permit_truth_sources_2026_audit.json"
    audit_json["outputs"] = {
        "audit_csv": str(audit_csv_path),
        "audit_json": str(audit_json_path),
    }
    with audit_json_path.open("w", encoding="utf-8") as handle:
        json.dump(audit_json, handle, indent=2)

    print(json.dumps(audit_json, indent=2))


if __name__ == "__main__":
    main()
