from __future__ import annotations

import csv
import json
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
SOURCE_2022_2024 = (
    REPO_ROOT
    / "data_truth"
    / "permit_overlay_truth"
    / "raw_sources"
    / "2022_2024_conservation_permits"
    / "source_files"
    / "2022_2024_conservation_permits_extracted.xlsx"
)
SOURCE_2025_2027 = REPO_ROOT / "data" / "conservation-permit-hunt-table-2025-27.csv"
OUT_DIR = REPO_ROOT / "data_truth" / "permit_overlay_truth" / "normalized"
DETAIL_CSV = OUT_DIR / "conservation_permit_cycle_rows_2022_2027.csv"
SPECIES_TREND_CSV = OUT_DIR / "conservation_permit_trends_by_species_2022_2027.csv"
GROUP_TREND_CSV = OUT_DIR / "conservation_permit_trends_by_group_2022_2027.csv"
SUMMARY_JSON = OUT_DIR / "conservation_permit_trends_2022_2027_summary.json"
REPORT_MD = REPO_ROOT / "processed_data" / "conservation_permit_trends_2022_2027.md"


SPECIES_FAMILY_MAP = {
    "Antlerless Elk": "Elk",
    "Bear": "Black Bear",
    "Bison": "Bison",
    "Black Bear": "Black Bear",
    "Buck Deer": "Deer",
    "Buck Pronghorn": "Pronghorn",
    "Bull Elk": "Elk",
    "Bull Moose": "Moose",
    "Cougar": "Cougar",
    "Deer": "Deer",
    "Desert Bighorn Sheep": "Desert Bighorn Sheep",
    "Elk": "Elk",
    "Moose": "Moose",
    "Mountain Goat": "Mountain Goat",
    "Pronghorn": "Pronghorn",
    "Rocky Mountain Bighorn Sheep": "Rocky Mountain Bighorn Sheep",
    "Turkey": "Turkey",
    "Wild Bearded Turkey": "Turkey",
}


DETAIL_FIELDS = [
    "cycle",
    "cycle_start_year",
    "cycle_end_year",
    "cycle_year_count",
    "source_file",
    "source_row_id",
    "group",
    "organization",
    "species_detail",
    "species_family",
    "area",
    "condition_or_weapon",
    "permit_count",
    "discontinued_flag",
    "source_page_or_row",
]

TREND_FIELDS = [
    "dimension",
    "key",
    "permits_2022_2024",
    "permits_2025_2027",
    "annual_avg_2022_2024",
    "annual_avg_2025_2027",
    "cycle_change",
    "annual_avg_change",
    "pct_change",
    "trend_direction",
]


def normalize_species(species: str) -> str:
    if species not in SPECIES_FAMILY_MAP:
        raise ValueError(f"Unmapped conservation species: {species!r}")
    return SPECIES_FAMILY_MAP[species]


def read_2022_2024_rows() -> list[dict[str, object]]:
    workbook = load_workbook(SOURCE_2022_2024, data_only=True)
    worksheet = workbook["Permit Data"]
    headers = [cell.value for cell in worksheet[1]]
    rows: list[dict[str, object]] = []

    for values in worksheet.iter_rows(min_row=2, values_only=True):
        raw = dict(zip(headers, values))
        species_detail = str(raw["species"])
        rows.append(
            {
                "cycle": "2022-2024",
                "cycle_start_year": 2022,
                "cycle_end_year": 2024,
                "cycle_year_count": 3,
                "source_file": "2022_2024_conservation_permits_extracted.xlsx",
                "source_row_id": raw["row_id"],
                "group": raw["group"],
                "organization": raw["organization"],
                "species_detail": species_detail,
                "species_family": normalize_species(species_detail),
                "area": raw["area"],
                "condition_or_weapon": raw["condition"],
                "permit_count": 1,
                "discontinued_flag": str(raw["condition"]).lower() == "discontinued",
                "source_page_or_row": raw["source_page"],
            }
        )
    return rows


def read_2025_2027_rows() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    with SOURCE_2025_2027.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for raw in reader:
            species_detail = raw["species"]
            rows.append(
                {
                    "cycle": "2025-2027",
                    "cycle_start_year": 2025,
                    "cycle_end_year": 2027,
                    "cycle_year_count": 3,
                    "source_file": "conservation-permit-hunt-table-2025-27.csv",
                    "source_row_id": raw["huntCode"],
                    "group": raw["huntClass"],
                    "organization": raw["huntClass"],
                    "species_detail": species_detail,
                    "species_family": normalize_species(species_detail),
                    "area": raw["area"],
                    "condition_or_weapon": raw["conditions"] or raw["weapon"],
                    "permit_count": int(float(raw["permitCount"] or 0)),
                    "discontinued_flag": False,
                    "source_page_or_row": raw["sourceRowStart"],
                }
            )
    return rows


def write_csv(path: Path, rows: list[dict[str, object]], fields: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def aggregate(rows: list[dict[str, object]], key_field: str) -> dict[str, dict[str, float]]:
    result: dict[str, dict[str, float]] = defaultdict(lambda: {"2022-2024": 0.0, "2025-2027": 0.0})
    for row in rows:
        result[str(row[key_field])][str(row["cycle"])] += float(row["permit_count"])
    return result


def trend_rows(rows: list[dict[str, object]], dimension: str, key_field: str) -> list[dict[str, object]]:
    trends = []
    for key, values in sorted(aggregate(rows, key_field).items()):
        old = values["2022-2024"]
        new = values["2025-2027"]
        change = new - old
        pct = "" if old == 0 else round((change / old) * 100, 2)
        if change > 0:
            direction = "UP"
        elif change < 0:
            direction = "DOWN"
        else:
            direction = "FLAT"
        trends.append(
            {
                "dimension": dimension,
                "key": key,
                "permits_2022_2024": int(old) if old.is_integer() else old,
                "permits_2025_2027": int(new) if new.is_integer() else new,
                "annual_avg_2022_2024": round(old / 3, 2),
                "annual_avg_2025_2027": round(new / 3, 2),
                "cycle_change": int(change) if change.is_integer() else change,
                "annual_avg_change": round(change / 3, 2),
                "pct_change": pct,
                "trend_direction": direction,
            }
        )
    return trends


def validate(detail_rows: list[dict[str, object]], species_trends: list[dict[str, object]], group_trends: list[dict[str, object]]) -> list[str]:
    failures: list[str] = []
    total_2022 = sum(int(row["permit_count"]) for row in detail_rows if row["cycle"] == "2022-2024")
    total_2025 = sum(int(row["permit_count"]) for row in detail_rows if row["cycle"] == "2025-2027")
    discontinued_2022 = sum(1 for row in detail_rows if row["cycle"] == "2022-2024" and row["discontinued_flag"])

    if total_2022 != 318:
        failures.append(f"Expected 318 2022-2024 permit rows, found {total_2022}.")
    if total_2025 != 336:
        failures.append(f"Expected 336 2025-2027 permits, found {total_2025}.")
    if discontinued_2022 != 6:
        failures.append(f"Expected 6 discontinued 2022-2024 rows, found {discontinued_2022}.")
    if sum(int(row["permits_2022_2024"]) for row in species_trends) != total_2022:
        failures.append("Species trend 2022-2024 total does not reconcile.")
    if sum(int(row["permits_2025_2027"]) for row in species_trends) != total_2025:
        failures.append("Species trend 2025-2027 total does not reconcile.")
    if sum(int(row["permits_2022_2024"]) for row in group_trends) != total_2022:
        failures.append("Group trend 2022-2024 total does not reconcile.")
    if sum(int(row["permits_2025_2027"]) for row in group_trends) != total_2025:
        failures.append("Group trend 2025-2027 total does not reconcile.")
    return failures


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    REPORT_MD.parent.mkdir(parents=True, exist_ok=True)

    detail_rows = read_2022_2024_rows() + read_2025_2027_rows()
    species_trends = trend_rows(detail_rows, "species_family", "species_family")
    group_trends = trend_rows(detail_rows, "group", "group")
    failures = validate(detail_rows, species_trends, group_trends)

    total_2022 = sum(int(row["permit_count"]) for row in detail_rows if row["cycle"] == "2022-2024")
    total_2025 = sum(int(row["permit_count"]) for row in detail_rows if row["cycle"] == "2025-2027")
    summary = {
        "classification": "CONSERVATION_PERMIT_TREND_MODEL_2022_2027",
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "method": "deterministic_cycle_to_cycle_permit_count_trend",
        "source_2022_2024": str(SOURCE_2022_2024.relative_to(REPO_ROOT)).replace("\\", "/"),
        "source_2025_2027": str(SOURCE_2025_2027.relative_to(REPO_ROOT)).replace("\\", "/"),
        "guardrail": "TREND_MODEL_ONLY_DO_NOT_USE_AS_DRAW_ODDS_OR_HARVEST_RESULTS",
        "cycle_totals": {
            "2022_2024": total_2022,
            "2025_2027": total_2025,
            "cycle_change": total_2025 - total_2022,
            "pct_change": round(((total_2025 - total_2022) / total_2022) * 100, 2),
            "annual_avg_2022_2024": round(total_2022 / 3, 2),
            "annual_avg_2025_2027": round(total_2025 / 3, 2),
        },
        "largest_species_increases": sorted(species_trends, key=lambda row: row["cycle_change"], reverse=True)[:5],
        "largest_species_decreases": sorted(species_trends, key=lambda row: row["cycle_change"])[:5],
        "largest_group_increases": sorted(group_trends, key=lambda row: row["cycle_change"], reverse=True)[:5],
        "largest_group_decreases": sorted(group_trends, key=lambda row: row["cycle_change"])[:5],
        "validation": {
            "status": "FAIL" if failures else "PASS",
            "failures": failures,
            "detail_row_count": len(detail_rows),
            "species_trend_row_count": len(species_trends),
            "group_trend_row_count": len(group_trends),
        },
    }

    write_csv(DETAIL_CSV, detail_rows, DETAIL_FIELDS)
    write_csv(SPECIES_TREND_CSV, species_trends, TREND_FIELDS)
    write_csv(GROUP_TREND_CSV, group_trends, TREND_FIELDS)
    SUMMARY_JSON.write_text(json.dumps(summary, indent=2) + "\n", encoding="utf-8")

    REPORT_MD.write_text(
        "\n".join(
            [
                "# Conservation Permit Trends 2022-2027",
                "",
                "Deterministic trend model comparing two Utah conservation permit cycles.",
                "",
                f"- 2022-2024 permits: `{total_2022}`",
                f"- 2025-2027 permits: `{total_2025}`",
                f"- Cycle change: `{total_2025 - total_2022}`",
                f"- Percent change: `{summary['cycle_totals']['pct_change']}%`",
                f"- Annual average change: `{round((total_2025 - total_2022) / 3, 2)}` permits/year",
                "",
                "## Guardrail",
                "",
                "`TREND_MODEL_ONLY_DO_NOT_USE_AS_DRAW_ODDS_OR_HARVEST_RESULTS`",
                "",
                "## Outputs",
                "",
                f"- `{DETAIL_CSV.relative_to(REPO_ROOT)}`",
                f"- `{SPECIES_TREND_CSV.relative_to(REPO_ROOT)}`",
                f"- `{GROUP_TREND_CSV.relative_to(REPO_ROOT)}`",
                f"- `{SUMMARY_JSON.relative_to(REPO_ROOT)}`",
                "",
            ]
        ),
        encoding="utf-8",
    )

    print(json.dumps(summary["validation"], indent=2))
    return 0 if not failures else 1


if __name__ == "__main__":
    raise SystemExit(main())
