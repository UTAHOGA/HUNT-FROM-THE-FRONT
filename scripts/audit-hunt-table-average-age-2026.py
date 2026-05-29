from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\tyler\Desktop\GitHub\HUNT-BUILDER")
XLSX_DIR = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "XLXS"
PDF_DIR = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "PDF'S"
AGE_SOURCE = ROOT / "processed_data" / "harvest_age_features_by_hunt_code_latest.csv"
AUDIT_CSV = ROOT / "processed_data" / "audits" / "hunt_tables_2026_average_age_cell_audit.csv"
AUDIT_JSON = ROOT / "processed_data" / "audits" / "hunt_tables_2026_average_age_cell_audit.json"

AGE_HEADER_ALIASES = [
    "Average Age Harvested (previous hunting season)",
    "Average Age Harvested",
    "Average Harvest Age",
    "Avg Age Harvested",
]


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def norm_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())


def norm_code(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", "", clean(value).upper())


def numeric_display(value: Any) -> str:
    try:
        number = float(clean(value))
    except ValueError:
        return ""
    if number <= 0:
        return ""
    return f"{number:.1f}".rstrip("0").rstrip(".")


def load_pass_age_lookup() -> dict[str, dict[str, str]]:
    lookup: dict[str, dict[str, str]] = {}
    if not AGE_SOURCE.exists():
        return lookup

    with AGE_SOURCE.open(newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            if clean(row.get("review_status")).upper() != "PASS":
                continue
            code = norm_code(row.get("hunt_code"))
            age = numeric_display(row.get("average_harvest_age"))
            if not code or not age:
                continue
            year_text = clean(row.get("reported_hunt_year"))
            try:
                year = int(float(year_text))
            except ValueError:
                year = 0
            prior_year = int(lookup.get(code, {}).get("reported_hunt_year", "0") or 0)
            if year >= prior_year:
                lookup[code] = {
                    "average_harvest_age": age,
                    "reported_hunt_year": str(year) if year else "",
                    "species": clean(row.get("species")),
                    "source_file": clean(row.get("source_file")),
                    "source_package_file": clean(row.get("source_package_file")),
                    "review_status": clean(row.get("review_status")),
                }
    return lookup


def find_header_row(ws) -> int:
    best_row = 1
    best_score = -1
    for row in range(1, min(ws.max_row, 14) + 1):
        values = [clean(cell.value).lower() for cell in ws[row]]
        joined = " | ".join(v for v in values if v)
        score = sum(1 for v in values if v)
        if "hunt_code" in joined or "hunt code" in joined:
            score += 20
        if "average age" in joined:
            score += 10
        if score > best_score:
            best_score = score
            best_row = row
    return best_row


def header_map(ws, header_row: int) -> dict[str, int]:
    return {
        norm_header(ws.cell(header_row, col).value): col
        for col in range(1, ws.max_column + 1)
        if clean(ws.cell(header_row, col).value)
    }


def fill_workbooks(age_lookup: dict[str, dict[str, str]]) -> dict[str, Any]:
    filled_by_file: Counter[str] = Counter()
    xlsx_files = sorted(path for path in XLSX_DIR.glob("*.xlsx") if not path.name.startswith("~$"))

    for workbook_path in xlsx_files:
        wb = load_workbook(workbook_path)
        ws = wb[wb.sheetnames[0]]
        header_row = find_header_row(ws)
        headers = header_map(ws, header_row)
        code_col = headers.get(norm_header("hunt_code")) or headers.get(norm_header("Hunt Code"))
        age_col = None
        for alias in AGE_HEADER_ALIASES:
            age_col = age_col or headers.get(norm_header(alias))

        if not code_col or not age_col:
            wb.close()
            continue

        changed = False
        for row_number in range(header_row + 1, ws.max_row + 1):
            code = norm_code(ws.cell(row_number, code_col).value)
            if not code:
                continue
            existing = clean(ws.cell(row_number, age_col).value)
            hard_age = age_lookup.get(code, {}).get("average_harvest_age", "")
            if existing or not hard_age:
                continue
            ws.cell(row_number, age_col).value = hard_age
            filled_by_file[workbook_path.name] += 1
            changed = True

        if changed:
            wb.save(workbook_path)
        wb.close()

    return {
        "files_with_cells_filled": sum(1 for count in filled_by_file.values() if count),
        "cells_filled_this_run": sum(filled_by_file.values()),
        "filled_by_file": dict(sorted(filled_by_file.items())),
    }


def scan_workbooks(age_lookup: dict[str, dict[str, str]]) -> tuple[list[dict[str, str]], dict[str, Any]]:
    rows: list[dict[str, str]] = []
    status_counts: Counter[str] = Counter()
    file_counts: dict[str, Counter[str]] = defaultdict(Counter)
    species_counts: Counter[str] = Counter()

    xlsx_files = sorted(path for path in XLSX_DIR.glob("*.xlsx") if not path.name.startswith("~$"))
    pdf_files = sorted(path for path in PDF_DIR.glob("*.pdf"))
    pdf_names = {path.stem for path in pdf_files}

    for workbook_path in xlsx_files:
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        header_row = find_header_row(ws)
        headers = header_map(ws, header_row)
        code_col = headers.get(norm_header("hunt_code")) or headers.get(norm_header("Hunt Code"))
        name_col = headers.get(norm_header("hunt_name")) or headers.get(norm_header("Hunt Name"))
        species_col = headers.get(norm_header("species"))
        age_col = None
        for alias in AGE_HEADER_ALIASES:
            age_col = age_col or headers.get(norm_header(alias))

        if not code_col or not age_col:
            status = "BLOCK_MISSING_REQUIRED_COLUMNS"
            rows.append(
                {
                    "workbook": workbook_path.name,
                    "pdf_exists": str(workbook_path.stem in pdf_names).lower(),
                    "row_number": "",
                    "hunt_code": "",
                    "hunt_name": "",
                    "species": "",
                    "age_cell_value": "",
                    "hard_data_age": "",
                    "hard_data_reported_hunt_year": "",
                    "hard_data_source_file": "",
                    "status": status,
                }
            )
            status_counts[status] += 1
            file_counts[workbook_path.name][status] += 1
            wb.close()
            continue

        for row_number in range(header_row + 1, ws.max_row + 1):
            code = norm_code(ws.cell(row_number, code_col).value)
            if not code:
                continue
            age_value = clean(ws.cell(row_number, age_col).value)
            hard = age_lookup.get(code, {})
            hard_age = hard.get("average_harvest_age", "")
            if age_value and hard_age and numeric_display(age_value) == hard_age:
                status = "AGE_POPULATED_FROM_PASS_HARD_DATA"
            elif age_value and hard_age:
                status = "REVIEW_AGE_VALUE_DIFFERS_FROM_PASS_HARD_DATA"
            elif age_value and not hard_age:
                status = "REVIEW_AGE_POPULATED_WITHOUT_PASS_HARD_DATA"
            elif not age_value and hard_age:
                status = "REVIEW_BLANK_BUT_PASS_HARD_DATA_AVAILABLE"
            else:
                status = "AGE_BLANK_NO_PASS_NUMERIC_HARD_DATA"

            species = clean(ws.cell(row_number, species_col).value) if species_col else ""
            status_counts[status] += 1
            file_counts[workbook_path.name][status] += 1
            if status == "AGE_POPULATED_FROM_PASS_HARD_DATA":
                species_counts[species] += 1
            rows.append(
                {
                    "workbook": workbook_path.name,
                    "pdf_exists": str(workbook_path.stem in pdf_names).lower(),
                    "row_number": str(row_number),
                    "hunt_code": code,
                    "hunt_name": clean(ws.cell(row_number, name_col).value) if name_col else "",
                    "species": species,
                    "age_cell_value": age_value,
                    "hard_data_age": hard_age,
                    "hard_data_reported_hunt_year": hard.get("reported_hunt_year", ""),
                    "hard_data_source_file": hard.get("source_file", ""),
                    "status": status,
                }
            )
        wb.close()

    missing_pdfs = [path.name for path in xlsx_files if path.stem not in pdf_names]
    orphan_pdfs = [path.name for path in pdf_files if not (XLSX_DIR / f"{path.stem}.xlsx").exists()]
    summary = {
        "created_at": datetime.now(timezone.utc).isoformat(),
        "xlsx_files": len(xlsx_files),
        "pdf_files": len(pdf_files),
        "missing_pdf_companions": missing_pdfs,
        "orphan_pdfs_without_xlsx": orphan_pdfs,
        "hard_age_source": str(AGE_SOURCE.relative_to(ROOT)),
        "pass_numeric_hard_age_codes": len(age_lookup),
        "rows_audited": len(rows),
        "status_counts": dict(status_counts),
        "age_populated_from_pass_hard_data": status_counts["AGE_POPULATED_FROM_PASS_HARD_DATA"],
        "age_blank_no_pass_numeric_hard_data": status_counts["AGE_BLANK_NO_PASS_NUMERIC_HARD_DATA"],
        "review_blank_but_pass_hard_data_available": status_counts["REVIEW_BLANK_BUT_PASS_HARD_DATA_AVAILABLE"],
        "review_age_value_differs_from_pass_hard_data": status_counts["REVIEW_AGE_VALUE_DIFFERS_FROM_PASS_HARD_DATA"],
        "review_age_populated_without_pass_hard_data": status_counts["REVIEW_AGE_POPULATED_WITHOUT_PASS_HARD_DATA"],
        "file_counts": {file: dict(counts) for file, counts in sorted(file_counts.items())},
        "species_populated_counts": dict(species_counts),
        "outputs": {
            "audit_csv": str(AUDIT_CSV.relative_to(ROOT)),
            "audit_json": str(AUDIT_JSON.relative_to(ROOT)),
        },
    }
    return rows, summary


def main() -> None:
    age_lookup = load_pass_age_lookup()
    rows, summary = scan_workbooks(age_lookup)
    fill_summary = {
        "files_with_cells_filled": 0,
        "cells_filled_this_run": 0,
        "filled_by_file": {},
        "mode": "AUDIT_ONLY_NO_XLSX_EDITS",
    }
    summary["fill_summary"] = fill_summary

    AUDIT_CSV.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "workbook",
        "pdf_exists",
        "row_number",
        "hunt_code",
        "hunt_name",
        "species",
        "age_cell_value",
        "hard_data_age",
        "hard_data_reported_hunt_year",
        "hard_data_source_file",
        "status",
    ]
    with AUDIT_CSV.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    AUDIT_JSON.write_text(json.dumps(summary, indent=2) + "\n", encoding="utf-8")

    print(f"xlsx_files={summary['xlsx_files']}")
    print(f"pdf_files={summary['pdf_files']}")
    print(f"rows_audited={summary['rows_audited']}")
    print(f"cells_filled_this_run={fill_summary['cells_filled_this_run']}")
    print(f"files_with_cells_filled={fill_summary['files_with_cells_filled']}")
    print(f"age_populated_from_pass_hard_data={summary['age_populated_from_pass_hard_data']}")
    print(f"age_blank_no_pass_numeric_hard_data={summary['age_blank_no_pass_numeric_hard_data']}")
    print(f"review_blank_but_pass_hard_data_available={summary['review_blank_but_pass_hard_data_available']}")
    print(f"review_age_value_differs_from_pass_hard_data={summary['review_age_value_differs_from_pass_hard_data']}")
    print(f"review_age_populated_without_pass_hard_data={summary['review_age_populated_without_pass_hard_data']}")
    print(f"audit_csv={AUDIT_CSV}")
    print(f"audit_json={AUDIT_JSON}")


if __name__ == "__main__":
    main()
