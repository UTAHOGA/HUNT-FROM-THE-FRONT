from __future__ import annotations

import csv
from pathlib import Path
from typing import Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

ROOT = Path(r"C:\Users\tyler\Desktop\GitHub\HUNT-BUILDER")
TARGET_DIR = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "XLXS"
AUDIT_CSV = ROOT / "processed_data" / "audits" / "hunt_tables_2026_xlsx_reformat_audit.csv"


def find_header_row(ws) -> Optional[int]:
    probe_max = min(ws.max_row, 14)
    best_row = None
    best_score = -1
    for r in range(1, probe_max + 1):
        vals = [str(c.value).strip() if c.value is not None else "" for c in ws[r]]
        non = [v for v in vals if v]
        if not non:
            continue
        key = " | ".join(v.lower() for v in non)
        score = len(non)
        if "hunt_code" in key:
            score += 20
        if "hunt_name" in key:
            score += 12
        if "species" in key:
            score += 6
        if score > best_score:
            best_score = score
            best_row = r
    return best_row


def header_bounds(ws, header_row: int) -> Optional[Tuple[int, int]]:
    max_col = ws.max_column
    first_col = None
    last_col = None
    for c in range(1, max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None or str(v).strip() == "":
            continue
        if first_col is None:
            first_col = c
        last_col = c
    if first_col is None or last_col is None:
        return None
    return first_col, last_col


def find_last_data_row(ws, header_row: int, first_col: int, last_col: int) -> int:
    last = header_row
    for r in range(header_row + 1, ws.max_row + 1):
        has_val = False
        for c in range(first_col, last_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() != "":
                has_val = True
                break
        if has_val:
            last = r
    if last == header_row:
        last = header_row + 1
    return last


def reset_tables(ws) -> None:
    for name in list(ws.tables.keys()):
        del ws.tables[name]


def sanitize_header_row(ws, header_row: int, first_col: int, last_col: int) -> None:
    seen = set()
    for c in range(first_col, last_col + 1):
        raw = ws.cell(row=header_row, column=c).value
        text = "" if raw is None else str(raw).strip()
        if not text:
            text = f"column_{c}"
        base = text
        i = 2
        while text.lower() in seen:
            text = f"{base}_{i}"
            i += 1
        seen.add(text.lower())
        ws.cell(row=header_row, column=c).value = text


def apply_visual_polish(ws, header_row: int, first_col: int, last_col: int, last_row: int) -> None:
    # Use explicit fixed colors and borders so styling is identical across workbooks
    # regardless of workbook theme/template differences.
    title_fill = PatternFill(fill_type="solid", fgColor="EDE3D3")
    subtitle_fill = PatternFill(fill_type="solid", fgColor="F5EFE4")
    header_fill = PatternFill(fill_type="solid", fgColor="5B301B")
    odd_row_fill = PatternFill(fill_type="solid", fgColor="F7F1E8")
    even_row_fill = PatternFill(fill_type="solid", fgColor="EFE5D7")

    border_side = Side(style="thin", color="C58F61")
    grid_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    title_font = Font(name="Calibri", size=14, bold=True, color="2F1B0F")
    subtitle_font = Font(name="Calibri", size=10, italic=True, color="5B3A25")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10, color="000000")

    if header_row > 1:
        title = ws.cell(row=1, column=first_col)
        if title.value is not None:
            title.font = title_font
            title.fill = title_fill
            title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

    if header_row > 2:
        subtitle = ws.cell(row=2, column=first_col)
        if subtitle.value is not None:
            subtitle.font = subtitle_font
            subtitle.fill = subtitle_fill
            subtitle.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 40

    for c in range(first_col, last_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = grid_border

    for r in range(header_row + 1, last_row + 1):
        row_fill = odd_row_fill if ((r - header_row) % 2 == 1) else even_row_fill
        for c in range(first_col, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = data_font
            cell.fill = row_fill
            cell.border = grid_border

            # right-align obvious numeric columns; left-align text columns
            header_text = str(ws.cell(row=header_row, column=c).value or "").lower()
            if any(k in header_text for k in ["res", "non-res", "total", "percent", "avg", "age", "days"]):
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for c in range(first_col, last_col + 1):
        col_letter = get_column_letter(c)
        max_len = 0
        for r in range(header_row, min(last_row, header_row + 350) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        if max_len <= 10:
            width = 10
        elif max_len <= 20:
            width = 14
        elif max_len <= 35:
            width = 20
        else:
            width = 24
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = ws.cell(row=header_row + 1, column=first_col)
    ws.auto_filter.ref = f"{get_column_letter(first_col)}{header_row}:{get_column_letter(last_col)}{last_row}"

    # Print layout: landscape, narrow margins, centered, fit-to-page width.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
    else:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.35, bottom=0.35, header=0.2, footer=0.2)
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False
    ws.print_title_rows = f"${header_row}:${header_row}"


def process_file(path: Path, idx: int) -> dict:
    out = {
        "file": path.name,
        "sheet": "",
        "header_row": "",
        "first_col": "",
        "last_col": "",
        "last_row": "",
        "table_ref": "",
        "columns": "",
        "data_rows": "",
        "status": "",
        "error": "",
    }
    wb = None
    try:
        wb = load_workbook(path)
        ws = wb[wb.sheetnames[0]]
        out["sheet"] = ws.title

        hr = find_header_row(ws)
        if hr is None:
            out["status"] = "SKIP"
            out["error"] = "header_not_found"
            return out

        bounds = header_bounds(ws, hr)
        if bounds is None:
            out["status"] = "SKIP"
            out["error"] = "header_bounds_not_found"
            return out

        fc, lc = bounds
        lr = find_last_data_row(ws, hr, fc, lc)

        reset_tables(ws)
        sanitize_header_row(ws, hr, fc, lc)

        ref = f"{get_column_letter(fc)}{hr}:{get_column_letter(lc)}{lr}"
        tname = f"HUNT_TABLE_{idx:03d}"
        table = Table(displayName=tname, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

        apply_visual_polish(ws, hr, fc, lc, lr)

        wb.save(path)

        out.update(
            {
                "header_row": hr,
                "first_col": fc,
                "last_col": lc,
                "last_row": lr,
                "table_ref": ref,
                "columns": lc - fc + 1,
                "data_rows": max(0, lr - hr),
                "status": "OK",
            }
        )
        return out
    except Exception as exc:
        out["status"] = "ERROR"
        out["error"] = str(exc)
        return out
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def main() -> None:
    if not TARGET_DIR.exists():
        raise SystemExit(f"Missing target folder: {TARGET_DIR}")

    files = sorted([p for p in TARGET_DIR.glob("*.xlsx") if not p.name.startswith("~$")])
    if not files:
        raise SystemExit("No .xlsx files found")

    results = []
    for i, f in enumerate(files, start=1):
        res = process_file(f, i)
        results.append(res)
        print(f"{res['file']}: {res['status']} {res['table_ref']} {res['error']}")

    AUDIT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with AUDIT_CSV.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(
            fh,
            fieldnames=[
                "file",
                "sheet",
                "header_row",
                "first_col",
                "last_col",
                "last_row",
                "table_ref",
                "columns",
                "data_rows",
                "status",
                "error",
            ],
        )
        writer.writeheader()
        writer.writerows(results)

    ok = sum(1 for r in results if r["status"] == "OK")
    err = sum(1 for r in results if r["status"] == "ERROR")
    skip = sum(1 for r in results if r["status"] == "SKIP")
    print(f"DONE total={len(results)} ok={ok} skip={skip} error={err} audit={AUDIT_CSV}")


if __name__ == "__main__":
    main()
