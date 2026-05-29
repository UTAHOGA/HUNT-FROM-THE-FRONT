from __future__ import annotations

import csv
import json
import math
import re
import shutil
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.table import Table, TableStyleInfo
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, legal
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas


ROOT = Path(r"C:\Users\tyler\Desktop\GitHub\HUNT-BUILDER")
SOURCE_DIR = ROOT / "pipeline" / "RAW" / "hunt_unit_database" / "2026" / "formatted_xlsx"
XLSX_DIR = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "XLXS"
PDF_DIR = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "PDF'S"
AUDIT_CSV = ROOT / "processed_data" / "audits" / "hunt_tables_2026_clean_redraft_audit.csv"
PDF_MANIFEST = ROOT / "processed_data" / "hard_data_exports" / "hard_copy_pdf_manifest.web.json"
HARVEST_QUALITY = ROOT / "processed_data" / "harvest_quality_features_all_years_by_hunt_code.csv"
HARVEST_MASTER = ROOT / "processed_data" / "harvest_master.csv"

HEADERS = [
    "hunt_name",
    "hunt_code",
    "sex_type",
    "species",
    "weapon",
    "hunt_type",
    "season",
    "permits_2026_res",
    "permits_2026_nr",
    "permits_2026_total",
    "NOTES",
    "Harvest Prior Year",
    "Percent Harvest Success (previous hunting season)",
    "Average Age Harvested (previous hunting season)",
    "Avg Days Hunted (previous hunting season)",
]

COLUMN_WIDTHS = {
    "hunt_name": 24,
    "hunt_code": 11,
    "sex_type": 11,
    "species": 13,
    "weapon": 19,
    "hunt_type": 19,
    "season": 34,
    "permits_2026_res": 10,
    "permits_2026_nr": 10,
    "permits_2026_total": 10,
    "NOTES": 24,
    "Harvest Prior Year": 10,
    "Percent Harvest Success (previous hunting season)": 16,
    "Average Age Harvested (previous hunting season)": 16,
    "Avg Days Hunted (previous hunting season)": 16,
}

SKIP_SOURCE_FILES = {
    "DATABASE.xlsx",
    "hunt_master_canonical_2026_built.xlsx",
    "2026_utah_dwr_hunt_matrix.xlsx",
    # This workbook is two table blocks pasted side by side; the clean class files below cover the same scope.
    "2026_elk_bull_all 2.xlsx",
}

SOURCE_NAME_OVERRIDES = {
    "2026_ANTLERLESS_DEER_formatted": "2026 DEER ANTLERLESS DRAW",
    "2026_antlerless_elk_general_season": "2026 ELK ANTLERLESS GENERAL SEASON",
    "2026_bison_cow": "2026 BISON COW O.I.L",
    "2026_bison_hunter_choice": "2026 BISON HUNTER CHOICE O.I.L",
    "2026_black_bear": "2026 BLACK BEAR DRAW",
    "2026_cougar": "2026 COUGAR DRAW",
    "2026_deer_antlerless": "2026 DEER ANTLERLESS DRAW SUMMARY",
    "2026_deer_antlerless_cwmu": "2026 DEER ANTLERLESS CWMU",
    "2026_deer_archery_extended": "2026 DEER BUCK EXTENDED ARCHERY",
    "2026_deer_buck": "2026 DEER BUCK DRAW",
    "2026_deer_buck_cactus": "2026 DEER BUCK CACTUS BUCK",
    "2026_deer_buck_conservation": "2026 DEER BUCK CONSERVATION PERMIT",
    "2026_deer_buck_cwmu": "2026 DEER BUCK CWMU",
    "2026_deer_buck_limited_entry": "2026 DEER BUCK L.E",
    "2026_deer_buck_limited_entry_management_buck": "2026 DEER BUCK L.E MANAGEMENT BUCK",
    "2026_deer_buck_limited_entry_private_lands_only": "2026 DEER BUCK L.E PRIVATE LANDS ONLY",
    "2026_deer_buck_premium_limited_entry": "2026 DEER BUCK P.L.E",
    "2026_deer_buck_statewide": "2026 DEER BUCK STATEWIDE",
    "2026_deer_general_season": "2026 DEER BUCK GENERAL SEASON",
    "2026_deer_general_season_private_lands": "2026 DEER BUCK GENERAL SEASON PRIVATE LANDS",
    "2026_deer_hunter_choice": "2026 DEER HUNTER CHOICE",
    "2026_desert_bighorn_ram": "2026 DESERT BIGHORN SHEEP RAM O.I.L",
    "2026_elk_antlerless": "2026 ELK ANTLERLESS DRAW",
    "2026_elk_antlerless_conservation": "2026 ELK ANTLERLESS CONSERVATION PERMIT",
    "2026_elk_antlerless_CWMU": "2026 ELK ANTLERLESS CWMU",
    "2026_elk_antlerless_privatelandsonly": "2026 ELK ANTLERLESS PRIVATE LANDS ONLY",
    "2026_elk_archery_extended": "2026 ELK BULL EXTENDED ARCHERY",
    "2026_elk_bull_all": "2026 ELK BULL ALL HUNTS",
    "2026_elk_bull_conservation_permit": "2026 ELK BULL CONSERVATION PERMIT",
    "2026_elk_bull_cwmu": "2026 ELK BULL CWMU",
    "2026_elk_bull_general_anybull": "2026 ELK BULL GENERAL ANY BULL",
    "2026_elk_bull_limited_entry": "2026 ELK BULL L.E",
    "2026_elk_general_anybull_youth": "2026 ELK BULL YOUTH GENERAL ANY BULL",
    "2026_elk_general_archery": "2026 ELK BULL GENERAL ARCHERY",
    "2026_elk_general_spikeonly": "2026 ELK BULL GENERAL SPIKE ONLY",
    "2026_elk_limitedentry_maturebull": "2026 ELK BULL L.E MATURE BULL",
    "2026_elk_limitedentry_maturebull_privatelands": "2026 ELK BULL L.E MATURE BULL PRIVATE LANDS",
    "2026_elk_limitedentry_statewide": "2026 ELK BULL STATEWIDE",
    "2026_goat_hunter_choice ": "2026 MOUNTAIN GOAT HUNTER CHOICE O.I.L",
    "2026_moose_bull": "2026 MOOSE BULL O.I.L",
    "2026_moose_cow": "2026 MOOSE COW O.I.L",
    "2026_sheep_desert_ram": "2026 DESERT BIGHORN SHEEP RAM O.I.L SUMMARY",
    "2026_sheep_rocky_mountain_ewe": "2026 ROCKY MOUNTAIN BIGHORN SHEEP EWE O.I.L",
    "2026_sheep_rocky_mountain_ram": "2026 ROCKY MOUNTAIN BIGHORN SHEEP RAM O.I.L",
    "2026_turkey_bearded": "2026 TURKEY BEARDED DRAW",
    "2026_turkey_either_sex": "2026 TURKEY EITHER SEX DRAW",
}

PREFERRED_COLLISIONS = {
    "2026 COUGAR DRAW": "2026 COUGAR DRAW",
    "2026 DEER ANTLERLESS DRAW": "2026 DEER ANTLERLESS DRAW",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def norm_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean_text(value).lower())


def clean_stem(stem: str) -> str:
    if stem in SOURCE_NAME_OVERRIDES:
        return SOURCE_NAME_OVERRIDES[stem]
    text = re.sub(r"\s+", " ", stem.replace("_", " ")).strip()
    text = re.sub(r"\s+\.", "", text).strip()
    return text.upper()


def find_header_row(ws) -> int:
    best_row = 1
    best_score = -1
    for row in range(1, min(ws.max_row, 12) + 1):
        vals = [clean_text(ws.cell(row=row, column=col).value) for col in range(1, ws.max_column + 1)]
        joined = " ".join(v.lower() for v in vals if v)
        score = sum(1 for v in vals if v)
        if "hunt code" in joined or "hunt_code" in joined:
            score += 20
        if "hunt name" in joined or "hunt_name" in joined:
            score += 20
        if "species" in joined:
            score += 6
        if "season" in joined:
            score += 4
        if score > best_score:
            best_score = score
            best_row = row
    return best_row


def find_header_row_values(values: list[list[Any]]) -> int:
    best_row = 0
    best_score = -1
    for row_i, row in enumerate(values[:12]):
        joined = " ".join(clean_text(v).lower() for v in row if clean_text(v))
        score = sum(1 for v in row if clean_text(v))
        if "hunt code" in joined or "hunt_code" in joined:
            score += 20
        if "hunt name" in joined or "hunt_name" in joined:
            score += 20
        if "species" in joined:
            score += 6
        if "season" in joined:
            score += 4
        if score > best_score:
            best_score = score
            best_row = row_i
    return best_row


def header_index(ws, header_row: int) -> dict[str, int]:
    index: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = norm_key(ws.cell(row=header_row, column=col).value)
        if key and key not in index:
            index[key] = col
    return index


def header_index_values(row: list[Any]) -> dict[str, int]:
    index: dict[str, int] = {}
    for col_i, value in enumerate(row):
        key = norm_key(value)
        if key and key not in index:
            index[key] = col_i
    return index


def get_by_alias(ws, row: int, index: dict[str, int], aliases: list[str]) -> str:
    for alias in aliases:
        col = index.get(norm_key(alias))
        if not col:
            continue
        value = clean_text(ws.cell(row=row, column=col).value)
        if value:
            return value
    return ""


def get_value(row: list[Any], index: dict[str, int], aliases: list[str]) -> str:
    for alias in aliases:
        col = index.get(norm_key(alias))
        if col is None or col >= len(row):
            continue
        value = clean_text(row[col])
        if value:
            return value
    return ""


def load_harvest_lookup() -> dict[str, dict[str, str]]:
    lookup: dict[str, dict[str, str]] = {}

    def keep(code: str, item: dict[str, str], year: int) -> None:
        code = code.strip().upper()
        if not code:
            return
        prior_year = int(lookup.get(code, {}).get("_year", "0") or 0)
        if year >= prior_year:
            item["_year"] = str(year)
            lookup[code] = item

    if HARVEST_QUALITY.exists():
        with HARVEST_QUALITY.open(newline="", encoding="utf-8-sig") as fh:
            for row in csv.DictReader(fh):
                code = clean_text(row.get("hunt_code")).upper()
                year_text = clean_text(row.get("reported_hunt_year"))
                try:
                    year = int(float(year_text))
                except ValueError:
                    year = 0
                if year < 2025:
                    continue
                item = {
                    "Harvest Prior Year": str(year) if year else "",
                    "Percent Harvest Success (previous hunting season)": clean_text(row.get("percent_success")),
                    "Average Age Harvested (previous hunting season)": clean_text(row.get("average_age")),
                    "Avg Days Hunted (previous hunting season)": clean_text(row.get("average_days")),
                }
                if any(item.values()):
                    keep(code, item, year)

    if HARVEST_MASTER.exists():
        with HARVEST_MASTER.open(newline="", encoding="utf-8-sig") as fh:
            for row in csv.DictReader(fh):
                code = clean_text(row.get("hunt_code")).upper()
                year_text = clean_text(row.get("year"))
                try:
                    year = int(float(year_text))
                except ValueError:
                    year = 0
                if year < 2025:
                    continue
                current = lookup.get(code, {})
                item = {
                    "Harvest Prior Year": str(year) if year else current.get("Harvest Prior Year", ""),
                    "Percent Harvest Success (previous hunting season)": clean_text(row.get("percent_success"))
                    or current.get("Percent Harvest Success (previous hunting season)", ""),
                    "Average Age Harvested (previous hunting season)": current.get(
                        "Average Age Harvested (previous hunting season)", ""
                    ),
                    "Avg Days Hunted (previous hunting season)": clean_text(row.get("avg_days"))
                    or current.get("Avg Days Hunted (previous hunting season)", ""),
                }
                if any(item.values()):
                    keep(code, item, year)

    for item in lookup.values():
        item.pop("_year", None)
    return lookup


def extract_rows(path: Path, harvest_lookup: dict[str, dict[str, str]]) -> tuple[str, list[list[str]], int]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    values = [list(row) for row in ws.iter_rows(values_only=True)]
    header_i = find_header_row_values(values)
    header_row = header_i + 1
    index = header_index_values(values[header_i] if values else [])
    rows: list[list[str]] = []

    for source_row in values[header_i + 1 :]:
        record = {
            "hunt_name": get_value(source_row, index, ["hunt_name", "Hunt Name", "boundary_name", "Boundary Name"]),
            "hunt_code": get_value(source_row, index, ["hunt_code", "Hunt Code"]),
            "sex_type": get_value(source_row, index, ["sex_type", "Sex", "Sex Type"]),
            "species": get_value(source_row, index, ["species", "Species"]),
            "weapon": get_value(source_row, index, ["weapon", "Weapon"]),
            "hunt_type": get_value(source_row, index, ["hunt_type", "Hunt Type"]),
            "season": get_value(source_row, index, ["season", "Season"]),
            "permits_2026_res": get_value(source_row, index, ["permits_2026_res", "Res", "Resident"]),
            "permits_2026_nr": get_value(source_row, index, ["permits_2026_nr", "Non-Res", "Nonresident", "NR"]),
            "permits_2026_total": get_value(source_row, index, ["permits_2026_total", "Total"]),
            "NOTES": get_value(source_row, index, ["NOTES", "Notes", "description", "Description"]),
            "Harvest Prior Year": get_value(source_row, index, ["Harvest Prior Year"]),
            "Percent Harvest Success (previous hunting season)": get_value(
                source_row,
                index,
                ["Percent Harvest Success (previous hunting season)", "Percent Harvest Success"],
            ),
            "Average Age Harvested (previous hunting season)": get_value(
                source_row,
                index,
                [
                    "Average Age Harvested (previous hunting season)",
                    "Average Age Harvested",
                    "Average Harvest Age",
                    "Avg Age Harvested",
                ],
            ),
            "Avg Days Hunted (previous hunting season)": get_value(
                source_row,
                index,
                ["Avg Days Hunted (previous hunting season)", "Avg Days Hunted", "Average Days Hunted"],
            ),
        }
        harvest = harvest_lookup.get(record["hunt_code"].upper(), {})
        for field in [
            "Harvest Prior Year",
            "Percent Harvest Success (previous hunting season)",
            "Average Age Harvested (previous hunting season)",
            "Avg Days Hunted (previous hunting season)",
        ]:
            if not record[field] and harvest.get(field):
                record[field] = harvest[field]
        out = [record[header] for header in HEADERS]
        if any(out):
            rows.append(out)
    wb.close()
    return clean_stem(path.stem), rows, header_row


def style_workbook(ws, row_count: int) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="5B301B")
    odd_fill = PatternFill(fill_type="solid", fgColor="F7F1E8")
    even_fill = PatternFill(fill_type="solid", fgColor="EFE5D7")
    border_side = Side(style="thin", color="C58F61")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTHS[header]

    for row in range(2, row_count + 2):
        fill = odd_fill if row % 2 == 0 else even_fill
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = border
            cell.font = Font(name="Calibri", size=8, color="000000")
            if col in {8, 9, 10, 12, 13, 14, 15}:
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 28

    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(2, row_count + 1)}"
    table_ref = f"A1:{get_column_letter(len(HEADERS))}{max(2, row_count + 1)}"
    table = Table(displayName="HUNT_TABLE", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.15, right=0.15, top=0.25, bottom=0.25, header=0.1, footer=0.1)
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = "$1:$1"


def write_workbook(path: Path, rows: list[list[str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Hunt Table"
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    style_workbook(ws, len(rows))
    wb.save(path)


def wrap_cell(text: str, width: float, font_size: int, max_lines: int) -> list[str]:
    text = clean_text(text)
    if not text:
        return [""]
    max_chars = max(4, int(width / (font_size * 0.48)))
    words = text.split(" ")
    lines: list[str] = []
    line = ""
    for word in words:
        candidate = word if not line else f"{line} {word}"
        if len(candidate) <= max_chars:
            line = candidate
        else:
            if line:
                lines.append(line)
            line = word[:max_chars]
        if len(lines) >= max_lines:
            break
    if line and len(lines) < max_lines:
        lines.append(line)
    if len(lines) == max_lines and len(" ".join(words)) > len(" ".join(lines)):
        lines[-1] = lines[-1][: max(1, max_chars - 3)] + "..."
    return lines or [""]


def draw_pdf(path: Path, title: str, rows: list[list[str]]) -> None:
    page_w, page_h = landscape(legal)
    margin_x = 0.18 * inch
    margin_y = 0.22 * inch
    usable_w = page_w - (2 * margin_x)
    y_start = page_h - margin_y
    col_weights = [2.15, 0.9, 0.9, 1.05, 1.55, 1.55, 2.45, 0.75, 0.75, 0.8, 1.65, 0.75, 1.18, 1.18, 1.18]
    total_weight = sum(col_weights)
    col_widths = [usable_w * (w / total_weight) for w in col_weights]
    header_h = 0.42 * inch
    row_h = 0.34 * inch
    rows_per_page = max(1, math.floor((page_h - (2 * margin_y) - header_h - 0.22 * inch) / row_h))

    c = canvas.Canvas(str(path), pagesize=landscape(legal))

    def draw_header() -> None:
        c.setFont("Helvetica-Bold", 10)
        c.setFillColor(colors.HexColor("#2F1B0F"))
        c.drawCentredString(page_w / 2, y_start, title)
        y = y_start - 0.2 * inch
        x = margin_x
        c.setFillColor(colors.HexColor("#5B301B"))
        c.rect(margin_x, y - header_h, usable_w, header_h, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 5.9)
        for idx, header in enumerate(HEADERS):
            w = col_widths[idx]
            for line_i, line in enumerate(wrap_cell(header, w - 4, 5.9, 3)[:3]):
                c.drawString(x + 2, y - 9 - (line_i * 7), line)
            c.setStrokeColor(colors.HexColor("#C58F61"))
            c.rect(x, y - header_h, w, header_h, stroke=1, fill=0)
            x += w

    row_index = 0
    page = 0
    while row_index < len(rows) or page == 0:
        page += 1
        draw_header()
        y = y_start - 0.2 * inch - header_h
        page_rows = rows[row_index : row_index + rows_per_page]
        c.setFont("Helvetica", 5.8)
        for local_i, row in enumerate(page_rows):
            y -= row_h
            fill = colors.HexColor("#F7F1E8") if local_i % 2 == 0 else colors.HexColor("#EFE5D7")
            c.setFillColor(fill)
            c.rect(margin_x, y, usable_w, row_h, fill=1, stroke=0)
            x = margin_x
            c.setFillColor(colors.black)
            for idx, value in enumerate(row):
                w = col_widths[idx]
                lines = wrap_cell(value, w - 4, 5.8, 2)
                for line_i, line in enumerate(lines[:2]):
                    c.drawString(x + 2, y + row_h - 9 - (line_i * 7), line)
                c.setStrokeColor(colors.HexColor("#C58F61"))
                c.rect(x, y, w, row_h, stroke=1, fill=0)
                x += w
        row_index += len(page_rows)
        c.setFont("Helvetica", 7)
        c.setFillColor(colors.HexColor("#5B3A25"))
        c.drawRightString(page_w - margin_x, margin_y * 0.55, f"Page {page}")
        if row_index < len(rows):
            c.showPage()
    c.save()


def update_manifest(outputs: list[dict[str, Any]]) -> None:
    if not PDF_MANIFEST.exists():
        return
    items = json.loads(PDF_MANIFEST.read_text(encoding="utf-8-sig"))
    if not isinstance(items, list):
        return

    cleaned = [
        item
        for item in items
        if not (
            isinstance(item, dict)
            and str(item.get("href", "")).replace("\\", "/").startswith("./processed_data/hard_data_exports/hunt_tables/2026/")
        )
    ]
    base_order = 9000
    for idx, output in enumerate(outputs, start=1):
        title = output["name"]
        pdf_href = f"./processed_data/hard_data_exports/hunt_tables/2026/PDF'S/{title}.pdf"
        xlsx_href = f"./processed_data/hard_data_exports/hunt_tables/2026/XLXS/{title}.xlsx"
        cleaned.append(
            {
                "group": "hunt_tables",
                "type": "pdf",
                "year": "2026",
                "title": title,
                "subtitle": "Clean public display hunt table PDF.",
                "href": pdf_href,
                "companion_type": "xlsx",
                "companion_href": xlsx_href,
                "sort_order": base_order + idx,
            }
        )
    PDF_MANIFEST.write_text(json.dumps(cleaned, indent=2) + "\n", encoding="utf-8")


def main() -> None:
    if not SOURCE_DIR.exists():
        raise SystemExit(f"Missing source folder: {SOURCE_DIR}")
    source_files = sorted(
        [
            p
            for p in SOURCE_DIR.glob("*.xlsx")
            if not p.name.startswith("~$") and p.name not in SKIP_SOURCE_FILES
        ]
    )
    if not source_files:
        raise SystemExit("No source XLSX files found")

    harvest_lookup = load_harvest_lookup()
    records: list[dict[str, Any]] = []
    used_names: set[str] = set()
    for path in source_files:
        name, rows, header_row = extract_rows(path, harvest_lookup)
        if name in used_names:
            if name in PREFERRED_COLLISIONS:
                name = f"{name} SUMMARY"
            else:
                i = 2
                while f"{name} {i}" in used_names:
                    i += 1
                name = f"{name} {i}"
        used_names.add(name)
        records.append({"source": path.name, "name": name, "rows": rows, "header_row": header_row})

    staging_xlsx = XLSX_DIR.parent / "_XLXS_CLEAN_BUILD"
    staging_pdf = PDF_DIR.parent / "_PDF_CLEAN_BUILD"
    for folder in [staging_xlsx, staging_pdf]:
        if folder.exists():
            shutil.rmtree(folder)
        folder.mkdir(parents=True, exist_ok=True)

    for record in records:
        write_workbook(staging_xlsx / f"{record['name']}.xlsx", record["rows"])
        draw_pdf(staging_pdf / f"{record['name']}.pdf", record["name"], record["rows"])

    for folder in [XLSX_DIR, PDF_DIR]:
        folder.mkdir(parents=True, exist_ok=True)
        for file in folder.iterdir():
            if file.is_file() and file.suffix.lower() in {".xlsx", ".pdf"} and not file.name.startswith("~$"):
                file.unlink()

    for src in staging_xlsx.glob("*.xlsx"):
        shutil.move(str(src), str(XLSX_DIR / src.name))
    for src in staging_pdf.glob("*.pdf"):
        shutil.move(str(src), str(PDF_DIR / src.name))
    shutil.rmtree(staging_xlsx)
    shutil.rmtree(staging_pdf)

    update_manifest(records)

    AUDIT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with AUDIT_CSV.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=["source_file", "clean_file_stem", "source_header_row", "rows_written"])
        writer.writeheader()
        for record in records:
            writer.writerow(
                {
                    "source_file": record["source"],
                    "clean_file_stem": record["name"],
                    "source_header_row": record["header_row"],
                    "rows_written": len(record["rows"]),
                }
            )

    print(f"source_files={len(source_files)}")
    print(f"clean_xlsx={len(list(XLSX_DIR.glob('*.xlsx')))}")
    print(f"clean_pdf={len(list(PDF_DIR.glob('*.pdf')))}")
    print(f"harvest_lookup_codes={len(harvest_lookup)}")
    print(f"audit={AUDIT_CSV}")


if __name__ == "__main__":
    main()
