import csv
import json
from pathlib import Path
from copy import copy

from openpyxl import load_workbook


REPO = Path(__file__).resolve().parents[1]
HARVEST_FEATURES = REPO / "processed_data" / "harvest_quality_features_all_years_by_hunt_code.csv"
HARVEST_MASTER = REPO / "processed_data" / "harvest_master.csv"
HISTORICAL_CROSSWALK = REPO / "processed_data" / "current_to_historical_hunt_code_crosswalk_2026.csv"
TARGET_MODEL_YEAR = 2026

TARGET_DIRS = [
    REPO / "pipeline" / "RAW" / "hunt_unit_database" / "2026" / "formatted_tables",
    REPO / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026",
    REPO / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "XLXS",
]

HUNT_CODE_HEADERS = {
    "hunt code",
    "hunt_code",
    "huntcode",
    "hunt no",
    "hunt_no",
    "permit number",
    "permit_number",
}
SPECIES_HEADERS = {"species", "animal", "game"}
HUNT_NAME_HEADERS = {"hunt name", "hunt_name", "unit", "hunt unit", "area", "unit name"}
WEAPON_HEADERS = {"weapon", "weapon type", "weapon_type", "method", "method of take"}

COL_PRIOR_YEAR = "Harvest Prior Year"
COL_SUCCESS = "Percent Harvest Success (previous hunting season)"
COL_AVG_AGE = "Average Age Harvested (previous hunting season)"
COL_AVG_DAYS = "Avg Days Hunted (previous hunting season)"
OUTPUT_HEADERS = [COL_PRIOR_YEAR, COL_SUCCESS, COL_AVG_AGE, COL_AVG_DAYS]
HEADER_ALIASES = {
    COL_PRIOR_YEAR: {"harvest prior year"},
    COL_SUCCESS: {"harvest success (prior year %)", "percent harvest success (previous hunting season)"},
    COL_AVG_AGE: {"average harvest age (prior year)", "average age harvested (previous hunting season)"},
    COL_AVG_DAYS: {"average days hunted (prior year)", "avg days hunted (previous hunting season)"},
}


def normalize_code(value: str) -> str:
    return "".join(ch for ch in str(value or "").upper().strip() if ch.isalnum())


def normalize_text(value: str) -> str:
    return " ".join(
        "".join(ch if ch.isalnum() else " " for ch in str(value or "").lower()).split()
    )


def normalize_species(value: str) -> str:
    text = normalize_text(value)
    if "deer" in text:
        return "deer"
    if "elk" in text:
        return "elk"
    if "pronghorn" in text:
        return "pronghorn"
    if "moose" in text:
        return "moose"
    if "goat" in text:
        return "mountain goat"
    if "bear" in text:
        return "black bear"
    if "bison" in text:
        return "bison"
    if "sheep" in text:
        return "sheep"
    if "turkey" in text:
        return "turkey"
    return text


def normalize_weapon(value: str) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    if "any legal" in text:
        return "any legal weapon"
    if "archery" in text:
        return "archery"
    if "muzzle" in text:
        return "muzzleloader"
    if "rifle" in text:
        return "rifle"
    if "shotgun" in text:
        return "shotgun"
    if "crossbow" in text:
        return "crossbow"
    if "hunter choice" in text:
        return "hunter choice"
    return text


def normalize_hunt_name(value: str) -> str:
    text = normalize_text(value)
    stop_words = {
        "limited",
        "entry",
        "general",
        "season",
        "private",
        "lands",
        "only",
        "permit",
        "permits",
        "conservation",
        "expo",
        "statewide",
        "draw",
        "bonus",
        "preference",
        "mature",
        "bull",
        "buck",
        "cow",
        "doe",
    }
    tokens = [t for t in text.split() if t not in stop_words]
    return " ".join(tokens).strip()


def build_triplet_key(species: str, hunt_name: str, weapon: str) -> str:
    return "|".join(
        [
            normalize_species(species),
            normalize_hunt_name(hunt_name),
            normalize_weapon(weapon),
        ]
    )


def norm_header(value: str) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def to_int(value):
    try:
        return int(float(str(value).strip()))
    except Exception:
        return None


def to_float(value):
    try:
        return float(str(value).strip())
    except Exception:
        return None


def format_number(value):
    if value is None:
        return ""
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.2f}".rstrip("0").rstrip(".")


def choose_best(existing, candidate):
    if not existing:
        return candidate

    ex_rank = existing["rank"]
    ca_rank = candidate["rank"]
    if ca_rank > ex_rank:
        return candidate
    if ca_rank < ex_rank:
        return existing

    ex_year = existing.get("reported_year") or 0
    ca_year = candidate.get("reported_year") or 0
    if ca_year > ex_year:
        return candidate

    return existing


def get_col_index(header_cells, accepted_headers):
    for idx, value in enumerate(header_cells, start=1):
        if norm_header(value) in accepted_headers:
            return idx
    return None


def parse_code_list(text: str):
    parts = []
    for token in str(text or "").replace("|", ";").replace(",", ";").split(";"):
        code = normalize_code(token)
        if code:
            parts.append(code)
    seen = set()
    ordered = []
    for code in parts:
        if code not in seen:
            seen.add(code)
            ordered.append(code)
    return ordered


def load_crosswalk_candidates():
    mapping = {}
    if not HISTORICAL_CROSSWALK.exists():
        return mapping
    with HISTORICAL_CROSSWALK.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            current = normalize_code(row.get("current_hunt_code") or row.get("hunt_code") or "")
            if not current:
                continue
            candidates = []
            candidates.extend(parse_code_list(row.get("historical_hunt_code")))
            candidates.extend(parse_code_list(row.get("candidate_historical_codes")))
            if candidates:
                mapping[current] = candidates
    return mapping


def load_harvest_lookup():
    lookup = {}
    feature_lookup = {}
    triplet_candidates = {}

    if HARVEST_MASTER.exists():
        with HARVEST_MASTER.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                code = normalize_code(row.get("hunt_code"))
                if not code:
                    continue
                reported_year = to_int(row.get("year"))
                prediction_year = to_int(row.get("prediction_year"))
                if prediction_year == TARGET_MODEL_YEAR:
                    rank = 5
                elif reported_year == TARGET_MODEL_YEAR - 1:
                    rank = 4
                elif reported_year:
                    rank = 2
                else:
                    rank = 1
                candidate = {
                    "rank": rank,
                    "reported_year": reported_year,
                    "success_pct": to_float(row.get("percent_success")),
                    "avg_age": None,
                    "avg_days": to_float(row.get("avg_days")),
                }
                lookup[code] = choose_best(lookup.get(code), candidate)

    if HARVEST_FEATURES.exists():
        with HARVEST_FEATURES.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                code = normalize_code(row.get("hunt_code"))
                if not code:
                    continue

                reported_year = to_int(row.get("reported_hunt_year"))
                model_target_year = to_int(row.get("model_target_year"))

                if model_target_year == TARGET_MODEL_YEAR:
                    rank = 3
                elif reported_year == TARGET_MODEL_YEAR - 1:
                    rank = 2
                elif reported_year:
                    rank = 1
                else:
                    rank = 0

                candidate = {
                    "rank": rank,
                    "reported_year": reported_year,
                    "success_pct": to_float(row.get("percent_success")),
                    "avg_age": to_float(row.get("average_age")),
                    "avg_days": to_float(row.get("average_days")),
                }
                lookup[code] = choose_best(lookup.get(code), candidate)
                feature_lookup[code] = choose_best(feature_lookup.get(code), candidate)

                triplet_key = build_triplet_key(row.get("species"), row.get("hunt_name"), row.get("weapon"))
                if triplet_key != "||":
                    existing = triplet_candidates.get(triplet_key)
                    if not existing:
                        triplet_candidates[triplet_key] = {
                            "ages": set([candidate["avg_age"]]) if candidate.get("avg_age") is not None else set(),
                            "payload": candidate,
                            "count": 1,
                        }
                    else:
                        existing["count"] += 1
                        if candidate.get("avg_age") is not None:
                            existing["ages"].add(candidate["avg_age"])
                        existing["payload"] = choose_best(existing.get("payload"), candidate)

    # Preserve a stronger age signal from feature rows even when master rows win
    # overall ranking for success/day values.
    for code, feature_candidate in feature_lookup.items():
        if feature_candidate.get("avg_age") is None:
            continue
        current = lookup.get(code)
        if not current:
            lookup[code] = feature_candidate
            continue
        if current.get("avg_age") is None:
            current["avg_age"] = feature_candidate.get("avg_age")
            if not current.get("reported_year") and feature_candidate.get("reported_year"):
                current["reported_year"] = feature_candidate.get("reported_year")

    # If current hunt codes have no direct age row, bridge historical age via crosswalk.
    crosswalk = load_crosswalk_candidates()
    for current_code, historical_candidates in crosswalk.items():
        current_payload = lookup.get(current_code)
        if current_payload and current_payload.get("avg_age") is not None:
            continue
        best_historical = None
        for hist_code in historical_candidates:
            hist_payload = feature_lookup.get(hist_code)
            if not hist_payload or hist_payload.get("avg_age") is None:
                continue
            best_historical = choose_best(best_historical, hist_payload)
        if not best_historical:
            continue
        base = current_payload.copy() if current_payload else {
            "rank": best_historical.get("rank", 0),
            "reported_year": best_historical.get("reported_year"),
            "success_pct": None,
            "avg_days": None,
            "avg_age": None,
        }
        base["avg_age"] = best_historical.get("avg_age")
        if not base.get("reported_year"):
            base["reported_year"] = best_historical.get("reported_year")
        lookup[current_code] = base

    # Keep only conservative unique/consistent triplet matches:
    # - at least one age value
    # - all matched rows collapse to a single age value
    triplet_lookup = {}
    for key, entry in triplet_candidates.items():
        ages = {age for age in entry.get("ages", set()) if age is not None}
        if len(ages) != 1:
            continue
        payload = dict(entry.get("payload") or {})
        payload["avg_age"] = next(iter(ages))
        payload["match_method"] = "species_hunt_name_weapon"
        triplet_lookup[key] = payload

    return lookup, triplet_lookup


def find_hunt_code_col(header_cells):
    for idx, value in enumerate(header_cells, start=1):
        if norm_header(value) in HUNT_CODE_HEADERS:
            return idx
    return None


def find_header_row(ws, max_scan_rows=20):
    for row_idx in range(1, min(ws.max_row, max_scan_rows) + 1):
        row_vals = [ws.cell(row=row_idx, column=i).value for i in range(1, ws.max_column + 1)]
        if find_hunt_code_col(row_vals):
            return row_idx
    return None


def ensure_output_cols(ws, header_row):
    header_values = [ws.cell(row=header_row, column=i).value for i in range(1, ws.max_column + 1)]
    mapping = {}
    added_any = False
    for header in OUTPUT_HEADERS:
        found = None
        for idx, value in enumerate(header_values, start=1):
            existing = str(value or "").strip().lower()
            if existing == header.lower() or existing in HEADER_ALIASES.get(header, set()):
                found = idx
                ws.cell(row=header_row, column=idx).value = header
                break
        if found is None:
            found = ws.max_column + 1
            ws.cell(row=header_row, column=found).value = header
            added_any = True
            if found > 1:
                src = ws.cell(row=header_row, column=found - 1)
                dst = ws.cell(row=header_row, column=found)
                if src.has_style:
                    dst._style = copy(src._style)
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)
                dst.number_format = src.number_format
                dst.protection = copy(src.protection)
        mapping[header] = found
    return mapping, added_any


def update_sheet(ws, harvest_lookup, triplet_lookup):
    header_row = find_header_row(ws)
    if not header_row:
        return {"updated_rows": 0, "header_added": False}

    header_values = [ws.cell(row=header_row, column=i).value for i in range(1, ws.max_column + 1)]
    hunt_code_col = find_hunt_code_col(header_values)
    if not hunt_code_col:
        return {"updated_rows": 0, "header_added": False}
    species_col = get_col_index(header_values, SPECIES_HEADERS)
    hunt_name_col = get_col_index(header_values, HUNT_NAME_HEADERS)
    weapon_col = get_col_index(header_values, WEAPON_HEADERS)

    out_cols, header_added = ensure_output_cols(ws, header_row)
    updated = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        code = normalize_code(ws.cell(row=row_idx, column=hunt_code_col).value)
        if not code:
            continue

        payload = harvest_lookup.get(code)
        if not payload:
            continue

        # Conservative second pass: use species+hunt-name+weapon if code payload has no age.
        if payload.get("avg_age") is None and species_col and hunt_name_col and weapon_col:
            species = ws.cell(row=row_idx, column=species_col).value
            hunt_name = ws.cell(row=row_idx, column=hunt_name_col).value
            weapon = ws.cell(row=row_idx, column=weapon_col).value
            triplet_key = build_triplet_key(species, hunt_name, weapon)
            triplet_payload = triplet_lookup.get(triplet_key)
            if triplet_payload and triplet_payload.get("avg_age") is not None:
                payload = dict(payload)
                payload["avg_age"] = triplet_payload.get("avg_age")
                if not payload.get("reported_year"):
                    payload["reported_year"] = triplet_payload.get("reported_year")

        ws.cell(row=row_idx, column=out_cols[COL_PRIOR_YEAR]).value = payload.get("reported_year") or ""
        ws.cell(row=row_idx, column=out_cols[COL_SUCCESS]).value = format_number(payload.get("success_pct"))
        ws.cell(row=row_idx, column=out_cols[COL_AVG_AGE]).value = format_number(payload.get("avg_age"))
        ws.cell(row=row_idx, column=out_cols[COL_AVG_DAYS]).value = format_number(payload.get("avg_days"))
        updated += 1

    return {"updated_rows": updated, "header_added": header_added}


def process_workbook(file_path: Path, harvest_lookup, triplet_lookup):
    wb = load_workbook(file_path)
    updated_rows = 0
    touched_sheets = 0
    header_changes = 0
    for ws in wb.worksheets:
        result = update_sheet(ws, harvest_lookup, triplet_lookup)
        sheet_updates = result["updated_rows"]
        if result["header_added"]:
            header_changes += 1
        if sheet_updates or result["header_added"]:
            touched_sheets += 1
            updated_rows += sheet_updates

    if touched_sheets:
        wb.save(file_path)

    return {
        "file": str(file_path.relative_to(REPO)).replace("\\", "/"),
        "updated_rows": updated_rows,
        "touched_sheets": touched_sheets,
        "header_changes": header_changes,
    }


def main():
    harvest_lookup, triplet_lookup = load_harvest_lookup()
    reports = []
    seen = set()

    for root in TARGET_DIRS:
        if not root.exists():
            continue
        for workbook in sorted(root.rglob("*.xlsx")):
            key = str(workbook.resolve()).lower()
            if key in seen:
                continue
            seen.add(key)
            reports.append(process_workbook(workbook, harvest_lookup, triplet_lookup))

    touched = [r for r in reports if r["updated_rows"] > 0 or r.get("header_changes", 0) > 0]
    print(json.dumps({
        "ok": True,
        "harvest_lookup_codes": len(harvest_lookup),
        "triplet_lookup_keys": len(triplet_lookup),
        "workbooks_scanned": len(reports),
        "workbooks_updated": len(touched),
        "updates": touched,
    }, indent=2))


if __name__ == "__main__":
    main()
