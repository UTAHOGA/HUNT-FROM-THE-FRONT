from __future__ import annotations

import csv
import json
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


ROOT = Path(__file__).resolve().parents[1]
DATABASE_PATH = ROOT / "pipeline" / "RAW" / "hunt_unit_database" / "2026" / "csv" / "DATABASE.csv"
HARVEST_PATHS = [
    ROOT / "processed_data" / "harvest_quality_features_all_years_by_hunt_code.csv",
    ROOT / "pipeline" / "RAW" / "hunt_unit_database" / "2026" / "csv" / "harvest_quality_features_by_hunt_code_2025_for_2026.csv",
]
OUTPUT_DIR = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "CLEAN_XLXS_STAGED"
OUTPUT_PATH = OUTPUT_DIR / "MASTER.xlsx"
AUDIT_DIR = ROOT / "processed_data" / "audits"
AUDIT_PATH = AUDIT_DIR / "hunt_tables_2026_master_xlsx_validation.json"
SEASON_NOTE_FALLBACKS = {
    "BR1000": "Any open season on any open bear unit during the 2026 season",
    "BR1001": "Any open harvest-objective unit during the 2026 season until objective closure",
}

OUTPUT_COLUMNS = [
    "HUNT TYPE",
    "SPECIES",
    "SEX TYPE",
    "WEAPON",
    "HUNT CLASS",
    "SEASON",
    "HUNT NAME",
    "HUNT CODE",
    "2026 PERMITS RES",
    "2026 PERMITS NR",
    "2026 PERMITS TOTAL",
    "HARVEST YEAR",
    "2026 HARVEST SUCCESS",
    "2026 HARVEST AGE",
    "2026 HARVEST DAYS",
]

HEADER_FONT = Font(name="Arial", size=10, bold=False)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_ROW_HEIGHT = 36


def clean_text(value: object) -> str:
    return str(value or "").strip()


def clean_code(value: object) -> str:
    return "".join(ch for ch in clean_text(value).upper() if ch.isalnum())


def format_number(value: object) -> str:
    text = clean_text(value)
    if text == "":
        return ""
    try:
        number = float(text.replace(",", "").replace("%", ""))
    except ValueError:
        return text
    if number.is_integer():
        return str(int(number))
    return f"{number:.2f}".rstrip("0").rstrip(".")


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def resolve_harvest_path() -> Path | None:
    for candidate in HARVEST_PATHS:
        if candidate.exists():
            return candidate
    return None


def choose_permit_value(row: dict[str, str], fields: list[str]) -> str:
    for field in fields:
        value = clean_text(row.get(field))
        if value != "":
            return format_number(value)
    return ""


def normalize_weapon(value: object) -> str:
    text = clean_text(value).lower()
    if "archery" in text:
        return "ARCHERY"
    if "muzzle" in text:
        return "ML"
    return "A.L.W."


def is_expo(row: dict[str, str]) -> bool:
    haystack = " ".join(
        [
            clean_text(row.get("hunt_name")),
            clean_text(row.get("hunt_type")),
            clean_text(row.get("hunt_class")),
        ]
    ).lower()
    return "expo" in haystack


def is_cwmu(row: dict[str, str]) -> bool:
    haystack = " ".join(
        [
            clean_text(row.get("hunt_name")),
            clean_text(row.get("hunt_type")),
            clean_text(row.get("hunt_class")),
        ]
    ).lower()
    return "cwmu" in haystack


def is_conservation(row: dict[str, str]) -> bool:
    haystack = " ".join(
        [
            clean_text(row.get("hunt_name")),
            clean_text(row.get("hunt_type")),
            clean_text(row.get("hunt_class")),
        ]
    ).lower()
    return "conservation" in haystack


def is_spike(row: dict[str, str]) -> bool:
    haystack = " ".join(
        [
            clean_text(row.get("hunt_name")),
            clean_text(row.get("hunt_type")),
            clean_text(row.get("hunt_class")),
        ]
    ).lower()
    return "spike" in haystack


def is_general_bull(row: dict[str, str]) -> bool:
    hunt_type = clean_text(row.get("hunt_type")).lower()
    hunt_name = clean_text(row.get("hunt_name")).lower()
    return "general season - any bull" in hunt_type or "any bull" in hunt_name


def is_private(row: dict[str, str]) -> bool:
    haystack = " ".join(
        [
            clean_text(row.get("hunt_name")),
            clean_text(row.get("hunt_type")),
            clean_text(row.get("hunt_class")),
        ]
    ).lower()
    return "private land" in haystack or "private lands" in haystack


def is_youth(row: dict[str, str]) -> bool:
    haystack = " ".join(
        [
            clean_text(row.get("hunt_name")),
            clean_text(row.get("hunt_type")),
            clean_text(row.get("hunt_class")),
        ]
    ).lower()
    return "youth" in haystack


def is_sportsman(row: dict[str, str]) -> bool:
    hunt_type = clean_text(row.get("hunt_type")).lower()
    hunt_name = clean_text(row.get("hunt_name")).lower()
    draw_system = clean_text(row.get("draw_2026_system_type")).upper()
    return (
        draw_system == "SPORTSMAN_PERMIT"
        or "sportsman" in hunt_name
        or (hunt_type == "statewide" and "statewide permit" in hunt_name)
    )


def resolve_season(row: dict[str, str]) -> str:
    season = clean_text(row.get("season"))
    if season:
        return season
    return SEASON_NOTE_FALLBACKS.get(clean_code(row.get("hunt_code")), "")


def classify_hunt(row: dict[str, str]) -> tuple[str, str]:
    hunt_type = clean_text(row.get("hunt_type")).lower()

    if is_expo(row):
        return "L.E.", "EXPO"
    if is_cwmu(row):
        return "L.E.", "CWMU"
    if is_conservation(row):
        return "L.E.", "CONSERVATION"
    if is_spike(row):
        return "OTC", "SPIKE"
    if is_general_bull(row):
        return "OTC", "GENERAL BULL"
    if is_private(row):
        return "OTC", "PRIVATE"
    if is_youth(row):
        return "GENERAL", "YOUTH"
    if "pursuit" in hunt_type:
        return "PURSUIT", ""
    if "harvest objective" in hunt_type:
        return "HARVEST OBJECTIVE", ""
    if is_sportsman(row):
        return "SPORTSMAN", ""
    if "premium limited entry" in hunt_type:
        return "P.L.E.", ""
    if "once-in-a-lifetime" in hunt_type:
        return "O.I.L.", ""
    if "limited entry" in hunt_type:
        return "L.E.", ""
    if "general season" in hunt_type or "extended archery" in hunt_type or "fall management" in hunt_type or "spring general season" in hunt_type or hunt_type == "statewide" or hunt_type == "tribal":
        return "GENERAL", ""
    if hunt_type == "private lands only":
        return "OTC", "PRIVATE"
    if hunt_type == "spot and stalk":
        return "OTC", ""
    return "GENERAL", ""


def build_harvest_lookup() -> dict[str, dict[str, str]]:
    lookup: dict[str, dict[str, str]] = {}
    harvest_path = resolve_harvest_path()
    if not harvest_path:
        return lookup

    for row in read_csv_rows(harvest_path):
        if clean_text(row.get("model_target_year")) != "2026":
            continue
        hunt_code = clean_code(row.get("hunt_code"))
        if hunt_code == "":
            continue
        current = lookup.get(hunt_code)
        candidate_year = int(clean_text(row.get("reported_hunt_year")) or "0")
        current_year = int(clean_text(current.get("reported_hunt_year")) or "0") if current else -1
        if current is None or candidate_year >= current_year:
            lookup[hunt_code] = row
    return lookup


def build_rows() -> tuple[list[dict[str, str]], dict[str, int | list[str] | dict[str, int]]]:
    database_rows = read_csv_rows(DATABASE_PATH)
    harvest_lookup = build_harvest_lookup()
    output_rows: list[dict[str, str]] = []

    missing_season_codes: list[str] = []
    season_note_fallback_rows = 0
    total_only_rows = 0
    split_rows = 0
    harvest_joined_rows = 0
    classification_counts: Counter[str] = Counter()
    class_counts: Counter[str] = Counter()

    for source_row in database_rows:
        hunt_code = clean_code(source_row.get("hunt_code"))
        harvest_row = harvest_lookup.get(hunt_code, {})

        permit_res = choose_permit_value(
            source_row,
            ["permits_2026_draw_res", "permits_2026_res", "permit_allotment_2026_res"],
        )
        permit_nr = choose_permit_value(
            source_row,
            ["permits_2026_draw_nr", "permits_2026_nr", "permit_allotment_2026_nr"],
        )
        permit_total = choose_permit_value(
            source_row,
            ["permits_2026_draw_total", "permits_2026_total", "permit_allotment_2026_total"],
        )

        if permit_total and not permit_res and not permit_nr:
            total_only_rows += 1
        if permit_res or permit_nr:
            split_rows += 1

        resolved_season = resolve_season(source_row)
        if clean_text(source_row.get("season")) == "" and resolved_season != "":
            season_note_fallback_rows += 1
        if resolved_season == "":
            missing_season_codes.append(hunt_code)

        hunt_type_out, hunt_class_out = classify_hunt(source_row)
        classification_counts[hunt_type_out] += 1
        if hunt_class_out:
            class_counts[hunt_class_out] += 1

        if harvest_row:
            harvest_joined_rows += 1

        output_rows.append(
            {
                "HUNT TYPE": hunt_type_out,
                "SPECIES": clean_text(source_row.get("species")),
                "SEX TYPE": clean_text(source_row.get("sex_type")),
                "WEAPON": normalize_weapon(source_row.get("weapon")),
                "HUNT CLASS": hunt_class_out,
                "SEASON": resolved_season,
                "HUNT NAME": clean_text(source_row.get("hunt_name")),
                "HUNT CODE": hunt_code,
                "2026 PERMITS RES": permit_res,
                "2026 PERMITS NR": permit_nr,
                "2026 PERMITS TOTAL": permit_total,
                "HARVEST YEAR": format_number(harvest_row.get("reported_hunt_year", "")),
                "2026 HARVEST SUCCESS": format_number(harvest_row.get("percent_success", "")),
                "2026 HARVEST AGE": format_number(harvest_row.get("average_age", "")),
                "2026 HARVEST DAYS": format_number(harvest_row.get("average_days", "")),
            }
        )

    output_rows.sort(
        key=lambda row: (
            row["HUNT TYPE"],
            row["SPECIES"],
            row["SEX TYPE"],
            row["WEAPON"],
            row["HUNT CLASS"],
            row["HUNT NAME"],
            row["HUNT CODE"],
        )
    )

    audit = {
        "database_rows": len(database_rows),
        "master_rows": len(output_rows),
        "harvest_joined_rows": harvest_joined_rows,
        "missing_season_count": len(missing_season_codes),
        "missing_season_codes_sample": missing_season_codes[:25],
        "season_note_fallback_rows": season_note_fallback_rows,
        "permit_total_only_rows": total_only_rows,
        "permit_split_rows": split_rows,
        "hunt_type_counts": dict(sorted(classification_counts.items())),
        "hunt_class_counts": dict(sorted(class_counts.items())),
    }
    return output_rows, audit


def set_column_widths(worksheet) -> None:
    widths = {
        "A": 14,
        "B": 18,
        "C": 14,
        "D": 12,
        "E": 16,
        "F": 24,
        "G": 42,
        "H": 14,
        "I": 16,
        "J": 16,
        "K": 18,
        "L": 14,
        "M": 18,
        "N": 15,
        "O": 16,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def write_workbook(rows: list[dict[str, str]]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "MASTER"
    worksheet.append(OUTPUT_COLUMNS)

    for row in rows:
        worksheet.append([row.get(column, "") for column in OUTPUT_COLUMNS])

    for cell in worksheet[1]:
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = BODY_ALIGNMENT

    worksheet.row_dimensions[1].height = HEADER_ROW_HEIGHT
    worksheet.freeze_panes = "A2"
    set_column_widths(worksheet)
    workbook.save(OUTPUT_PATH)


def validate_workbook(expected_rows: list[dict[str, str]], audit: dict[str, object]) -> None:
    workbook = load_workbook(OUTPUT_PATH)
    worksheet = workbook["MASTER"]

    headers = [worksheet.cell(1, index).value for index in range(1, len(OUTPUT_COLUMNS) + 1)]
    if headers != OUTPUT_COLUMNS:
        raise ValueError(f"Header mismatch: {headers}")

    for header in headers:
        if header != str(header).upper():
            raise ValueError(f"Header is not all caps: {header}")
        if "_" in str(header):
            raise ValueError(f"Header contains underscore: {header}")

    if worksheet.row_dimensions[1].height < HEADER_ROW_HEIGHT:
        raise ValueError("Header row height is too small for wrapped text.")

    for cell in worksheet[1]:
        if cell.font.name != "Arial" or cell.font.size != 10:
            raise ValueError(f"Header font mismatch at {cell.coordinate}")
        if not cell.alignment.wrap_text:
            raise ValueError(f"Header wrap_text is not enabled at {cell.coordinate}")
        if cell.alignment.horizontal != "center":
            raise ValueError(f"Header alignment is not centered at {cell.coordinate}")

    actual_row_count = worksheet.max_row - 1
    if actual_row_count != len(expected_rows):
        raise ValueError(f"Row count mismatch: workbook={actual_row_count}, expected={len(expected_rows)}")

    workbook_codes = []
    for row_index in range(2, worksheet.max_row + 1):
        workbook_codes.append(clean_code(worksheet.cell(row_index, 8).value))
    if len(workbook_codes) != len(set(workbook_codes)):
        raise ValueError("Duplicate hunt codes detected in MASTER.xlsx")

    if audit["missing_season_count"] != 0:
        raise ValueError(f"Season values missing for {audit['missing_season_count']} rows.")

    validation = {
        "output_file": str(OUTPUT_PATH.relative_to(ROOT)).replace("\\", "/"),
        "headers_match": True,
        "row_count": actual_row_count,
        "duplicate_hunt_code_count": len(workbook_codes) - len(set(workbook_codes)),
        "missing_season_count": audit["missing_season_count"],
        "season_note_fallback_rows": audit["season_note_fallback_rows"],
        "permit_total_only_rows": audit["permit_total_only_rows"],
        "permit_split_rows": audit["permit_split_rows"],
        "harvest_joined_rows": audit["harvest_joined_rows"],
        "hunt_type_counts": audit["hunt_type_counts"],
        "hunt_class_counts": audit["hunt_class_counts"],
    }
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    AUDIT_PATH.write_text(json.dumps(validation, indent=2), encoding="utf-8")


def main() -> None:
    rows, audit = build_rows()
    write_workbook(rows)
    validate_workbook(rows, audit)
    print(f"Wrote {OUTPUT_PATH}")
    print(f"Wrote {AUDIT_PATH}")
    print(f"Rows: {len(rows)}")
    print(f"Harvest joins: {audit['harvest_joined_rows']}")
    print(f"Season note fallback rows: {audit['season_note_fallback_rows']}")
    print(f"Permit total-only rows: {audit['permit_total_only_rows']}")


if __name__ == "__main__":
    main()
