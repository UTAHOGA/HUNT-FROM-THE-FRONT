from __future__ import annotations

import argparse
import math
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table as PdfTable, TableStyle


DEFAULT_TITLE = "2026 Antlerless Deer Hunt Table"


def display_title(path: Path, provided: str | None) -> str:
    if provided:
        return provided
    return path.stem.replace("_", " ").replace("-", " ").title()


def clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].replace("-", "").isdigit():
        return text[:-2]
    return text


def normalized_header(value) -> str:
    text = clean_text(value)
    replacements = {
        "hunt_name": "Hunt Name",
        "hunt_code": "Hunt Code",
        "sex_type": "Sex",
        "species": "Species",
        "weapon": "Weapon",
        "hunt_type": "Hunt Type",
        "season": "Season",
        "permits_res_2026": "Res",
        "permits_non-res_2026": "Non-Res",
        "permits_total_2026": "Total",
        "permits_2026_res": "Res",
        "permits_2026_nr": "Non-Res",
        "permits_2026_total": "Total",
    }
    return replacements.get(text, text.replace("_", " ").replace("-", " ").title())


def is_note_header(value: str) -> bool:
    return clean_text(value).strip().lower() in {"note", "notes", "other"}


def is_resident_permit_header(value: str) -> bool:
    text = clean_text(value).strip().lower()
    return text in {"permits_res_2026", "permits_2026_res", "res"}


def is_nonresident_permit_header(value: str) -> bool:
    text = clean_text(value).strip().lower()
    return text in {"permits_non-res_2026", "permits_non_res_2026", "permits_2026_nr", "permits_2026_nres", "non-res", "non res"}


def is_total_permit_header(value: str) -> bool:
    text = clean_text(value).strip().lower()
    return text in {"permits_total_2026", "permits_2026_total", "total", "total_2026_permits"}


def has_meaningful_cell(values: Iterable[str]) -> bool:
    return any(clean_text(value) != "" for value in values)


def column_values(rows: list[list[str]], index: int) -> list[str]:
    return [row[index] if index < len(row) else "" for row in rows]


def prune_optional_columns(rows: list[list[str]]) -> list[list[str]]:
    """Drop print-only waste while preserving source values that are actually populated."""
    if not rows:
        return rows
    headers = rows[0]
    body = rows[1:]
    keep = [True] * len(headers)

    for idx, header in enumerate(headers):
        values = column_values(body, idx)
        if is_note_header(header) and not has_meaningful_cell(values):
            keep[idx] = False

    res_idx = next((idx for idx, header in enumerate(headers) if is_resident_permit_header(header)), None)
    nr_idx = next((idx for idx, header in enumerate(headers) if is_nonresident_permit_header(header)), None)
    total_idx = next((idx for idx, header in enumerate(headers) if is_total_permit_header(header)), None)
    if res_idx is not None and nr_idx is not None and total_idx is not None:
        res_values = column_values(body, res_idx)
        nr_values = column_values(body, nr_idx)
        total_values = column_values(body, total_idx)
        split_is_empty = not has_meaningful_cell(res_values) and not has_meaningful_cell(nr_values)
        split_is_all_zero = (
            has_meaningful_cell(total_values)
            and all(clean_text(value) in {"", "0"} for value in res_values)
            and all(clean_text(value) in {"", "0"} for value in nr_values)
        )
        if split_is_empty or split_is_all_zero:
            keep[res_idx] = False
            keep[nr_idx] = False

    return [[value for idx, value in enumerate(row) if idx < len(keep) and keep[idx]] for row in rows]


def rows_from_sheet(ws) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        values = [clean_text(value) for value in row]
        if any(values):
            rows.append(values)
    return rows


def best_column_widths(rows: list[list[str]]) -> dict[int, float]:
    widths: dict[int, float] = {}
    if not rows:
        return widths
    column_count = max(len(row) for row in rows)
    for col_idx in range(column_count):
        values = [row[col_idx] if col_idx < len(row) else "" for row in rows]
        longest = max((len(value) for value in values), default=8)
        if col_idx == 0:
            widths[col_idx + 1] = min(max(longest * 0.85, 18), 28)
        elif col_idx in (4, 6):
            widths[col_idx + 1] = min(max(longest * 0.75, 18), 34)
        elif col_idx >= 7:
            widths[col_idx + 1] = 10
        else:
            widths[col_idx + 1] = min(max(longest * 0.9, 10), 18)
    return widths


def format_workbook(input_path: Path, output_path: Path, title: str) -> list[list[str]]:
    wb = load_workbook(input_path)
    ws = wb.active
    raw_rows = rows_from_sheet(ws)
    if not raw_rows:
        raise ValueError(f"No visible table rows found in {input_path}")
    raw_rows = prune_optional_columns(raw_rows)

    headers = [normalized_header(value) for value in raw_rows[0]]
    body = raw_rows[1:]
    ws.delete_rows(1, ws.max_row)
    ws.delete_cols(1, ws.max_column)

    column_count = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
    title_cell = ws.cell(1, 1, title)
    title_cell.font = Font(name="Georgia", size=18, bold=True, color="3A1D06")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill("solid", fgColor="F6E9D6")
    ws.row_dimensions[1].height = 32

    ws.append([])
    ws.append(headers)
    for row in body:
        padded = row + [""] * (column_count - len(row))
        ws.append(padded[:column_count])

    header_row = 3
    first_data_row = 4
    last_row = ws.max_row
    last_col_letter = get_column_letter(column_count)

    orange = "D96F00"
    cream = "FFF7EC"
    dark = "2F2119"
    border_side = Side(style="thin", color="B8895F")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    for cell in ws[header_row]:
        cell.font = Font(name="Aptos Display", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=orange)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_idx in range(first_data_row, last_row + 1):
        fill = PatternFill("solid", fgColor=cream if row_idx % 2 == 0 else "FFFFFF")
        for col_idx in range(1, column_count + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.fill = fill
            cell.border = border
            cell.font = Font(name="Aptos", size=10, color=dark)
            if col_idx >= 8:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for col_idx, width in best_column_widths([headers, *body]).items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx in range(first_data_row, last_row + 1):
        longest = max(len(clean_text(ws.cell(row_idx, col_idx).value)) for col_idx in range(1, column_count + 1))
        ws.row_dimensions[row_idx].height = min(max(24, math.ceil(longest / 38) * 16), 54)

    table_ref = f"A3:{last_col_letter}{last_row}"
    table = Table(displayName="HuntTable2026", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = table_ref
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.45
    ws.page_margins.bottom = 0.35
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = "3:3"
    ws.oddHeader.center.text = title
    ws.oddHeader.center.size = 12
    ws.oddHeader.center.font = "Georgia,Bold"
    ws.oddFooter.center.text = "Page &P of &N"
    ws.print_area = f"A1:{last_col_letter}{last_row}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return [headers, *body]


def pdf_column_widths(page_width: float, column_count: int) -> list[float]:
    if column_count == 10:
        proportions = [1.85, 0.85, 0.95, 0.8, 1.65, 1.1, 1.85, 0.55, 0.7, 0.55]
    else:
        proportions = [1.0] * column_count
    total = sum(proportions)
    return [page_width * p / total for p in proportions]


def build_pdf(rows: list[list[str]], output_path: Path, title: str) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    page_size = landscape(letter)
    margin = 0.28 * inch
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=page_size,
        leftMargin=margin,
        rightMargin=margin,
        topMargin=0.35 * inch,
        bottomMargin=0.28 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "HuntTableTitle",
        parent=styles["Title"],
        fontName="Times-Bold",
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#3A1D06"),
        alignment=TA_CENTER,
        spaceAfter=8,
    )
    cell_style = ParagraphStyle(
        "HuntTableCell",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=7.3,
        leading=8.2,
        textColor=colors.HexColor("#2F2119"),
    )
    header_style = ParagraphStyle(
        "HuntTableHeader",
        parent=cell_style,
        fontName="Helvetica-Bold",
        textColor=colors.white,
        alignment=TA_CENTER,
    )

    table_data = []
    for row_idx, row in enumerate(rows):
        style = header_style if row_idx == 0 else cell_style
        table_data.append([Paragraph(clean_text(value), style) for value in row])

    available_width = page_size[0] - doc.leftMargin - doc.rightMargin
    col_widths = pdf_column_widths(available_width, len(rows[0]))
    table = PdfTable(table_data, repeatRows=1, colWidths=col_widths, hAlign="CENTER")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D96F00")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7.5),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#B8895F")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FFF7EC")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("ALIGN", (7, 1), (-1, -1), "CENTER"),
    ]))

    story = [Paragraph(title, title_style), Spacer(1, 0.06 * inch), table]
    doc.build(story)


def main() -> None:
    parser = argparse.ArgumentParser(description="Format a UOGA hunt table workbook and matching PDF.")
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--title", default=None)
    parser.add_argument("--output-dir", type=Path, default=None)
    args = parser.parse_args()

    input_path = args.input
    title = display_title(input_path, args.title or DEFAULT_TITLE)
    output_dir = args.output_dir or input_path.parent / "formatted"
    output_xlsx = output_dir / f"{input_path.stem}_formatted.xlsx"
    output_pdf = output_dir / f"{input_path.stem}_formatted.pdf"

    rows = format_workbook(input_path, output_xlsx, title)
    build_pdf(rows, output_pdf, title)
    print(f"XLSX: {output_xlsx}")
    print(f"PDF: {output_pdf}")


if __name__ == "__main__":
    main()
