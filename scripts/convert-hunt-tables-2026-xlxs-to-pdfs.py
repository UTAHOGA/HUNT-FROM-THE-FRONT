from __future__ import annotations

from datetime import datetime, timezone
import json
from pathlib import Path
import re
from typing import List, Tuple

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, legal
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

ROOT = Path(r"C:\Users\tyler\Desktop\GitHub\HUNT-BUILDER")
SRC_DIR = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "XLXS"
OUT_DIR = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "PDF'S"
PDF_MANIFEST = ROOT / "processed_data" / "hard_data_exports" / "hard_copy_pdf_manifest.web.json"


def norm(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).strip())


def find_header_row(ws) -> int:
    probe_max = min(ws.max_row, 14)
    best_r, best_score = 1, -1
    for r in range(1, probe_max + 1):
        vals = [norm(c.value) for c in ws[r]]
        non = [v for v in vals if v]
        if not non:
            continue
        key = " | ".join(v.lower() for v in non)
        score = len(non)
        if "hunt_code" in key:
            score += 20
        if "hunt_name" in key or "hunt name" in key:
            score += 12
        if "species" in key:
            score += 6
        if score > best_score:
            best_score = score
            best_r = r
    return best_r


def extract(ws, source_name: str) -> Tuple[str, str, List[str], List[List[str]]]:
    hr = find_header_row(ws)
    title = source_name
    subtitle = "2026 Utah hunt table with public permit and harvest-quality display fields."

    # bounds from header row
    first_col = None
    last_col = None
    for c in range(1, ws.max_column + 1):
        v = norm(ws.cell(row=hr, column=c).value)
        if v:
            if first_col is None:
                first_col = c
            last_col = c
    if first_col is None:
        first_col, last_col = 1, ws.max_column

    headers = [norm(ws.cell(row=hr, column=c).value) or f"Column {c}" for c in range(first_col, last_col + 1)]
    rows: List[List[str]] = []
    for r in range(hr + 1, ws.max_row + 1):
        vals = [norm(ws.cell(row=r, column=c).value) for c in range(first_col, last_col + 1)]
        if any(vals):
            rows.append(vals)

    return title, subtitle, headers, rows


def fit_widths(headers: List[str], rows: List[List[str]], avail_w: float) -> List[float]:
    weights = []
    sample = rows[:350]
    for i, h in enumerate(headers):
        m = len(h)
        for row in sample:
            if i < len(row):
                m = max(m, len(row[i]))
        m = max(8, min(48, m))
        hl = h.lower()
        if "hunt name" in hl:
            m = max(m, 26)
        if "season" in hl:
            m = max(m, 24)
        if "note" in hl:
            m = max(m, 22)
        weights.append(m)

    total = sum(weights)
    widths = [(w / total) * avail_w for w in weights]
    min_w = 0.48 * inch
    max_w = 2.1 * inch
    widths = [max(min_w, min(max_w, w)) for w in widths]
    scale = avail_w / sum(widths)
    return [w * scale for w in widths]


def render_pdf(xlsx: Path, pdf: Path) -> Tuple[int, int]:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    title, subtitle, headers, rows = extract(ws, xlsx.stem)
    wb.close()

    doc = SimpleDocTemplate(
        str(pdf),
        pagesize=landscape(legal),
        leftMargin=0.22 * inch,
        rightMargin=0.22 * inch,
        topMargin=0.28 * inch,
        bottomMargin=0.28 * inch,
    )

    styles = getSampleStyleSheet()
    st_title = ParagraphStyle("t", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=13, textColor=colors.HexColor("#2f1b0f"), spaceAfter=4)
    st_sub = ParagraphStyle("s", parent=styles["Normal"], fontName="Helvetica-Oblique", fontSize=8.5, textColor=colors.HexColor("#5b3a25"), spaceAfter=6)
    st_head = ParagraphStyle("head", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=6.9, leading=7.6, textColor=colors.white, alignment=1)
    st_cell = ParagraphStyle("cell", parent=styles["Normal"], fontName="Helvetica", fontSize=6.2, leading=7.1, textColor=colors.black)
    data = [[Paragraph(h, st_head) for h in headers]]
    data.extend([[Paragraph(v, st_cell) for v in row] for row in rows])

    widths = fit_widths(headers, rows, doc.width)
    tbl = Table(data, repeatRows=1, colWidths=widths, hAlign="CENTER")
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#4f2d1d")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,0), 8.1),
        ("LEADING", (0,0), (-1,0), 9.2),
        ("FONTNAME", (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE", (0,1), (-1,-1), 7.3),
        ("LEADING", (0,1), (-1,-1), 8.6),
        ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#c7b59f")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.HexColor("#fbf6ef"), colors.HexColor("#f3e8da")]),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("ALIGN", (0,0), (-1,-1), "LEFT"),
        ("LEFTPADDING", (0,0), (-1,-1), 3),
        ("RIGHTPADDING", (0,0), (-1,-1), 3),
        ("TOPPADDING", (0,0), (-1,-1), 2),
        ("BOTTOMPADDING", (0,0), (-1,-1), 2),
    ]))

    meta = f"Source workbook: {xlsx.name} | Rows: {len(rows)} | Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    story = [Paragraph(title, st_title)]
    if subtitle:
        story.append(Paragraph(subtitle, st_sub))
    story.append(Paragraph(meta, st_sub))
    story.append(Spacer(1, 0.06 * inch))
    story.append(tbl)

    doc.build(story)
    return len(headers), len(rows)


def update_manifest() -> None:
    if not PDF_MANIFEST.exists():
        return
    items = json.loads(PDF_MANIFEST.read_text(encoding="utf-8-sig"))
    cleaned = [
        item for item in items
        if "hunt_tables/2026" not in str(item.get("href", ""))
        and "hunt_tables/2026" not in str(item.get("companion_href", ""))
    ]
    for idx, pdf in enumerate(sorted(OUT_DIR.glob("*.pdf")), start=1):
        title = pdf.stem
        xlsx = SRC_DIR / f"{title}.xlsx"
        cleaned.append(
            {
                "group": "hunt_tables",
                "type": "pdf",
                "year": "2026",
                "title": title,
                "subtitle": "Clean public display hunt table PDF.",
                "href": f"./processed_data/hard_data_exports/hunt_tables/2026/PDF'S/{pdf.name}",
                "companion_type": "xlsx",
                "companion_href": f"./processed_data/hard_data_exports/hunt_tables/2026/XLXS/{xlsx.name}",
                "sort_order": 9000 + idx,
            }
        )
    PDF_MANIFEST.write_text(json.dumps(cleaned, indent=2) + "\n", encoding="utf-8")


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    for old in OUT_DIR.glob("*.pdf"):
        old.unlink()
    files = sorted([p for p in SRC_DIR.glob("*.xlsx") if not p.name.startswith("~$")])
    if not files:
        raise SystemExit("No xlsx files found")

    ok = 0
    for x in files:
        out = OUT_DIR / (x.stem + ".pdf")
        try:
            cols, rows = render_pdf(x, out)
            print(f"OK {x.name} -> {out.name} cols={cols} rows={rows}")
            ok += 1
        except Exception as e:
            print(f"ERR {x.name}: {e}")

    update_manifest()
    print(f"DONE total={len(files)} ok={ok} out={OUT_DIR}")
    print(f"MANIFEST {PDF_MANIFEST}")

if __name__ == '__main__':
    main()
