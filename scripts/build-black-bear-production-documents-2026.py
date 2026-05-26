"""Build production library documents for reviewed 2026 black bear permits."""

from __future__ import annotations

import csv
import hashlib
import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table as PdfTable, TableStyle


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "pipeline/RAW/hunt_unit_database/2026/csv/2026 Permits/2026 black bear permits reviewed res-nr-total.csv"
OUTPUT_DIR = ROOT / "processed_data/hard_data_exports/hunt_tables/2026"
XLSX_OUT = OUTPUT_DIR / "2026_BLACK_BEAR.xlsx"
PDF_OUT = OUTPUT_DIR / "2026_BLACK_BEAR.pdf"
VALIDATION_OUT = OUTPUT_DIR / "2026_BLACK_BEAR.validation.json"

TITLE = "2026 BLACK BEAR"
SOURCE_NOTE = (
    "Source: reviewed 2026 Utah DWR Hunt Planner black bear permits export. "
    "Resident, nonresident, and total columns are split for library use. "
    "Blank numeric cells mean no published numeric permit count in the reviewed source; "
    "BR7307 is 2026 code reuse: it is now the La Sal multiseason conservation package "
    "with 4 total permits and no published resident/nonresident split, while the older "
    "La Sal limited-entry multiseason history crosswalks to current BR7326."
)

FIELDS = [
    "hunt_name",
    "hunt_code",
    "sex_type",
    "species",
    "weapon",
    "hunt_type",
    "season",
    "permits_2026_res",
    "permits_2026_nr",
    "permits_2026_total",
]

HEADERS = [
    "Hunt Name",
    "Hunt Code",
    "Sex",
    "Species",
    "Weapon",
    "Hunt Type",
    "Season",
    "Res",
    "Non-Res",
    "Total",
]


def clean(value: object) -> str:
    return str(value or "").strip()


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def read_rows() -> list[dict[str, str]]:
    with SOURCE.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames != FIELDS:
            raise ValueError(f"Unexpected reviewed black bear export headers: {reader.fieldnames}")
        rows = [{field: clean(row.get(field, "")) for field in FIELDS} for row in reader]
    codes = [row["hunt_code"] for row in rows]
    duplicate_codes = sorted(code for code, count in Counter(codes).items() if count > 1)
    if duplicate_codes:
        raise ValueError(f"Duplicate hunt codes in reviewed black bear export: {duplicate_codes}")
    if len(rows) != 106:
        raise ValueError(f"Expected 106 reviewed black bear rows, found {len(rows)}")
    return rows


def numeric_or_blank(value: str) -> int | str:
    if value == "":
        return ""
    return int(value)


def write_xlsx(rows: list[dict[str, str]]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "2026_BLACK_BEAR"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    title_cell = ws.cell(1, 1, TITLE)
    title_cell.font = Font(name="Georgia", size=18, bold=True, color="3A1D06")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill("solid", fgColor="F6E9D6")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
    note_cell = ws.cell(2, 1, SOURCE_NOTE)
    note_cell.font = Font(name="Aptos", size=9, italic=True, color="5A4737")
    note_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[2].height = 34

    ws.append(HEADERS)
    for row in rows:
        ws.append(
            [
                row["hunt_name"],
                row["hunt_code"],
                row["sex_type"],
                row["species"],
                row["weapon"],
                row["hunt_type"],
                row["season"],
                numeric_or_blank(row["permits_2026_res"]),
                numeric_or_blank(row["permits_2026_nr"]),
                numeric_or_blank(row["permits_2026_total"]),
            ]
        )

    header_row = 3
    first_data_row = 4
    last_row = ws.max_row
    last_col_letter = get_column_letter(len(HEADERS))

    orange = "D96F00"
    cream = "FFF7EC"
    dark = "2F2119"
    border_side = Side(style="thin", color="B8895F")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    for cell in ws[header_row]:
        cell.font = Font(name="Aptos Display", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=orange)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_idx in range(first_data_row, last_row + 1):
        fill = PatternFill("solid", fgColor=cream if row_idx % 2 == 0 else "FFFFFF")
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.fill = fill
            cell.border = border
            cell.font = Font(name="Aptos", size=10, color=dark)
            cell.alignment = Alignment(
                horizontal="center" if col_idx >= 8 else "left",
                vertical="center",
                wrap_text=True,
            )

    widths = [28, 11, 13, 12, 22, 24, 42, 9, 9, 9]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row_idx in range(first_data_row, last_row + 1):
        ws.row_dimensions[row_idx].height = 38

    table = Table(displayName="BlackBearPermits2026", ref=f"A3:{last_col_letter}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{last_col_letter}{last_row}"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:3"
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.45
    ws.page_margins.bottom = 0.45

    wb.save(XLSX_OUT)


def paragraph(text: str, style: ParagraphStyle) -> Paragraph:
    return Paragraph(clean(text).replace("&", "&amp;"), style)


def write_pdf(rows: list[dict[str, str]]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "HuntTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#3A1D06"),
        spaceAfter=8,
    )
    note_style = ParagraphStyle(
        "SourceNote",
        parent=styles["BodyText"],
        fontName="Helvetica-Oblique",
        fontSize=7.4,
        leading=9,
        textColor=colors.HexColor("#5A4737"),
        spaceAfter=8,
    )
    cell_style = ParagraphStyle(
        "Cell",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=6.4,
        leading=7.4,
        textColor=colors.HexColor("#2F2119"),
    )
    header_style = ParagraphStyle(
        "Header",
        parent=cell_style,
        fontName="Helvetica-Bold",
        textColor=colors.white,
        alignment=TA_CENTER,
    )
    number_style = ParagraphStyle(
        "Number",
        parent=cell_style,
        alignment=TA_CENTER,
    )

    doc = SimpleDocTemplate(
        str(PDF_OUT),
        pagesize=landscape(letter),
        rightMargin=0.28 * inch,
        leftMargin=0.28 * inch,
        topMargin=0.32 * inch,
        bottomMargin=0.34 * inch,
        title=TITLE,
        author="HUNTS verified source pipeline",
    )

    table_data = [[paragraph(header, header_style) for header in HEADERS]]
    for row in rows:
        table_data.append(
            [
                paragraph(row["hunt_name"], cell_style),
                paragraph(row["hunt_code"], number_style),
                paragraph(row["sex_type"], cell_style),
                paragraph(row["species"], cell_style),
                paragraph(row["weapon"], cell_style),
                paragraph(row["hunt_type"], cell_style),
                paragraph(row["season"], cell_style),
                paragraph(row["permits_2026_res"], number_style),
                paragraph(row["permits_2026_nr"], number_style),
                paragraph(row["permits_2026_total"], number_style),
            ]
        )

    col_widths = [1.72, 0.66, 0.82, 0.74, 1.12, 1.35, 2.38, 0.42, 0.54, 0.44]
    table = PdfTable(
        table_data,
        repeatRows=1,
        colWidths=[width * inch for width in col_widths],
        splitByRow=True,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D96F00")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (1, 0), (1, -1), "CENTER"),
                ("ALIGN", (7, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B8895F")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FFF7EC")]),
                ("TOPPADDING", (0, 0), (-1, -1), 2.2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2.2),
                ("LEFTPADDING", (0, 0), (-1, -1), 2.0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2.0),
            ]
        )
    )

    story = [
        Paragraph(TITLE, title_style),
        Paragraph(SOURCE_NOTE, note_style),
        Spacer(1, 0.04 * inch),
        table,
    ]

    def add_page_number(canvas, document) -> None:
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#5A4737"))
        canvas.drawRightString(
            document.pagesize[0] - document.rightMargin,
            0.18 * inch,
            f"Page {document.page}",
        )
        canvas.restoreState()

    doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)


def row_status(row: dict[str, str]) -> str:
    if row["permits_2026_res"] or row["permits_2026_nr"]:
        return "FULL_SPLIT"
    if row["permits_2026_total"]:
        return "TOTAL_ONLY"
    return "NO_PUBLISHED_NUMERIC_PERMIT"


def write_validation(rows: list[dict[str, str]]) -> None:
    status_counts = Counter(row_status(row) for row in rows)
    total_numeric = sum(int(row["permits_2026_total"] or 0) for row in rows)
    by_code = {row["hunt_code"]: row for row in rows}
    validation = {
        "built_at_utc": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "source_file": str(SOURCE.relative_to(ROOT)).replace("\\", "/"),
        "source_sha256": sha256(SOURCE),
        "xlsx_file": str(XLSX_OUT.relative_to(ROOT)).replace("\\", "/"),
        "pdf_file": str(PDF_OUT.relative_to(ROOT)).replace("\\", "/"),
        "row_count": len(rows),
        "unique_hunt_code_count": len(by_code),
        "status_counts": dict(sorted(status_counts.items())),
        "numeric_total_permits": total_numeric,
        "checked_codes": {
            code: {
                "hunt_name": by_code[code]["hunt_name"],
                "res": by_code[code]["permits_2026_res"],
                "nr": by_code[code]["permits_2026_nr"],
                "total": by_code[code]["permits_2026_total"],
                "status": row_status(by_code[code]),
                "code_reuse_warning": (
                    "2026 BR7307 is code reuse: it is the La Sal multiseason conservation package; "
                    "older La Sal limited-entry multiseason history crosswalks to current BR7326."
                    if code == "BR7307"
                    else ""
                ),
            }
            for code in ["BR7004", "BR7210", "BR7211", "BR7307", "BR7317", "BR7326"]
        },
    }
    VALIDATION_OUT.write_text(json.dumps(validation, indent=2) + "\n", encoding="utf-8")


def main() -> None:
    rows = read_rows()
    write_xlsx(rows)
    write_pdf(rows)
    write_validation(rows)
    print(f"Wrote {XLSX_OUT.relative_to(ROOT)}")
    print(f"Wrote {PDF_OUT.relative_to(ROOT)}")
    print(f"Wrote {VALIDATION_OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
